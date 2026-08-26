"""استخراج فرمول هر ستون فایل اکسل شرکت، کنار عنوان همان ستون.

    python tools/dump_excel_formulas.py 1405.xlsx > docs/EXCEL-FORMULAS.md

چرا این ابزار هست: فرمول‌ها تنها جایی هستند که «روش واقعی شرکت» در آن نوشته
شده. هر بار که سؤالی دربارهٔ یک قلم پیش می‌آید، باید فایل را باز کرد، ستون را
پیدا کرد و فرمول را از نوار فرمول خواند. این اسکریپت همان کار را یک بار انجام
می‌دهد و نتیجه را متنی و قابل جست‌وجو می‌کند.

**فرمولِ غالب** گزارش می‌شود، نه فرمول یک سطر: در هر ستون ممکن است چند سطر
دست‌کاری شده باشند. اگر ستونی بیش از یک شکل فرمول داشته باشد، همه‌شان با
تعداد سطرهایشان می‌آیند — و همان استثناها معمولاً مهم‌ترین چیزی‌اند که باید
دیده شوند.
"""

import collections
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalise(formula, row):
    """شمارهٔ سطر را از فرمول برمی‌دارد تا فرمول دو سطر قابل مقایسه شود.

    `=K15-(L15+M15)` و `=K36-(L36+M36)` یک فرمول‌اند؛ بدون این جایگزینی، هر
    سطر یک «شکل» جدا شمرده می‌شد و گزارش بی‌فایده می‌شد.
    """
    text = re.sub(r"\s+", " ", str(formula)).strip()
    return re.sub(rf"(?<=[A-Z]){row}\b", "#", text)


def header_row(sheet):
    for row in range(1, 6):
        for col in range(1, sheet.max_column + 1):
            if clean(sheet.cell(row, col).value) == "نام ونام خانوادگي":
                return row
    return 1


def describe(sheet):
    head = header_row(sheet)
    rows = [
        r for r in range(head + 1, sheet.max_row + 1)
        if sheet.cell(r, 5).value
    ]
    out = []
    for col in range(1, sheet.max_column + 1):
        title = clean(sheet.cell(head, col).value)
        if not title:
            continue
        shapes = collections.Counter()
        constants = collections.Counter()
        for r in rows:
            value = sheet.cell(r, col).value
            if isinstance(value, str) and value.startswith("="):
                shapes[normalise(value, r)] += 1
            elif value not in (None, ""):
                constants[type(value).__name__] += 1
        out.append({
            "letter": get_column_letter(col),
            "title": title,
            "shapes": shapes.most_common(),
            "constant_rows": sum(constants.values()),
            "total_rows": len(rows),
        })
    return out, len(rows)


def main(path, only=None):
    book = openpyxl.load_workbook(path, data_only=False)
    print("# فرمول‌های فایل اکسل حقوق و دستمزد\n")
    print(f"مرجع: `{path.split(chr(92))[-1]}`\n")
    print("این سند خودکار ساخته شده است:\n")
    print("```bash\npython tools/dump_excel_formulas.py <فایل> > docs/EXCEL-FORMULAS.md\n```\n")
    print("در هر ستون، شمارهٔ سطر با `#` جایگزین شده تا فرمول سطرهای مختلف")
    print("قابل مقایسه باشد. اگر ستونی بیش از یک شکل دارد، یعنی چند سطرش")
    print("**دستی عوض شده‌اند** — و همان‌ها معمولاً مهم‌ترین نکتهٔ ستون‌اند.\n")

    for name in book.sheetnames:
        if only and only not in name:
            continue
        sheet = book[name]
        if sheet.max_row < 2:
            continue
        columns, people = describe(sheet)
        if not people:
            continue
        print(f"\n---\n\n## شیت «{name}» — {people} سطر\n")
        for info in columns:
            print(f"### `{info['letter']}` · {info['title']}\n")
            if not info["shapes"]:
                print(f"عدد یا متنِ ورودی — {info['constant_rows']} سطر مقدار دارد.\n")
                continue
            for shape, count in info["shapes"]:
                tag = "همهٔ سطرها" if count == people else f"{count} سطر"
                print(f"- **{tag}**\n")
                print(f"  ```\n  {shape}\n  ```\n")
            if info["constant_rows"]:
                print(f"- {info['constant_rows']} سطر به‌جای فرمول، عددِ ثابت دارند.\n")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(1)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
