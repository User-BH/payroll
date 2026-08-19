"""بازسازی مستقل محاسبات فایل اکسل حقوق شرکت و مقایسه با مقادیر خود فایل.

این اسکریپت مدرکِ سند `docs/EXCEL-1405.md` است: هر فرمولی که آنجا نوشته شده،
اینجا در پایتون پیاده شده و روی همهٔ سطرهای همهٔ ماه‌ها با عددِ ذخیره‌شدهٔ اکسل
مقایسه می‌شود. اگر ماه تازه‌ای به فایل اضافه شد، با همین اسکریپت می‌شود فهمید
روش محاسبه عوض شده است یا نه — به‌جای اینکه دوباره ستون‌به‌ستون خوانده شود.

    python tools/verify_excel_1405.py /path/to/1405.xlsx

خروجی: برای هر ماه، تعداد سلول بازسازی‌شده و تعداد مغایرت. مغایرت یعنی یا فرمولِ
آن سلول در فایل با عدد دستی جایگزین شده، یا روش محاسبه عوض شده است.
"""

import sys
from decimal import Decimal as D, ROUND_DOWN, ROUND_HALF_UP, ROUND_UP

import openpyxl

# ---------------------------------------------------------------- گرد کردن اکسل
# ROUND اکسل «نیم به بالا»ست نه «نیم به زوج» پایتون؛ تفاوتشان روی ریال دیده می‌شود.
rnd = lambda x: D(str(x)).quantize(D("1"), rounding=ROUND_HALF_UP)
rdown = lambda x: D(str(x)).quantize(D("1"), rounding=ROUND_DOWN)
rup = lambda x: D(str(x)).quantize(D("1"), rounding=ROUND_UP)

WAGE_FACTOR = D("1.45")        # ضریب افزایش مزد ۱۴۰۵
WAGE_FIXED = D("519549")       # مبلغ ثابت افزایش ۱۴۰۵
OT_FACTOR = D("1.4")           # ضریب اضافه‌کاری
MONTH_HOURS = D("220")         # ساعت کار ماهانه
OT_MINUTES = D("13200")        # ۲۲۰ ساعت × ۶۰
LATE_MINUTES = D("11440")      # ۲۶ روز × ۴۴۰ دقیقه — مبنای متفاوتِ کسر کار
LEAVE_DAY_MINUTES = D("440")
CARRY_CAP = D("9")             # سقف انتقال مرخصی از سال قبل

TAX_BRACKETS = [               # (سقف، نرخ، مالیات انباشتهٔ پله‌های قبل)
    (D("400000000"), D("0"), D("0")),
    (D("800000000"), D("0.10"), D("0")),
    (D("1000000000"), D("0.15"), D("40000000")),
    (D("1200000000"), D("0.20"), D("70000000")),
    (D("1400000000"), D("0.25"), D("140000000")),
]
TOP_RATE, TOP_BASE = D("0.30"), D("190000000")


def income_tax(base: D) -> D:
    previous = D("0")
    for ceiling, rate, accumulated in TAX_BRACKETS:
        if base <= ceiling:
            return (base - previous) * rate + accumulated
        previous = ceiling
    return (base - previous) * TOP_RATE + TOP_BASE


# ------------------------------------------------------------------ عنوان ستون‌ها
# نگاشت با «عنوان» انجام می‌شود نه با حرف ستون: از تیر ۱۴۰۵ یک ستون کسور اضافه
# شده و همهٔ حرف‌های بعدش یکی جابه‌جا شده‌اند.
COLUMNS = {
    "days": "روز های ماه", "absence": "غیبت", "unpaid": "بدون حقوق", "sick": "بیماری",
    "worked": "کارکرد واقعی ماهانه", "late_h": "تاخیر ساعت", "late_m": "تاخیر دقیقه",
    "wage_day": "مزدروزانه", "sen_day": "پایه سنوات روزانه", "wage_sen_day": "جمع مزد وسنوات روزانه",
    "salary_month": "حقوق ماهانه", "sen_month": "پایه سنوات ماهانه",
    "salary_sen_month": "جمع حقوق وسنوات ماهانه", "cum_sen_day": "پایه سنوات تجمیعی روزانه",
    "cum_sen_month": "پایه سنوات تجمیعی ماهانه", "salary_pay": "حقوق قابل پرداخت",
    "sen_pay": "سنوات قابل پرداخت", "sen_diff": "مابه التفاوت پایه سنوات تجمیعی قابل پرداخت",
    "housing_pay": "حق مسكن قابل پرداخت", "child_pay": "حق اولاد قابل پرداخت",
    "marriage_pay": "حق تاهل قابل پرداخت", "food_pay": "وجه بن قابل پرداخت",
    "ot_min": "اضافه کاری عادی دقیقه", "ot_hour": "اضافه کاری  عادی ساعت",
    "ot_amount": "مبلغ اضافه كاري کل", "leave_pay": "وجه مرخصی", "end_pay": "پایانکار قابل پرداخت",
    "eid_pay": "عیدی وپاداش قابل پرداخت", "mission_days": "مدت ماموریت", "mission_amount": "مبلغ ماموریت",
    "comm_reserve": "ذخیره پورسانت فروش", "comm1_pay": "پورسانت یک قابل پرداخت",
    "comm2_pay": "پورسانت قابل پرداخت دو", "comm3_pay": "پورسانت قابل پرداخت سه",
    "prev_claim": "طلب ماه قبل", "gross": "جمع پرداختيها فیش", "adv_bank": "علی الحساب واریزی",
    "adv_cash": "علی الحساب نقدی", "adv_total": "جمع علی الحساب", "tax_base": "ماخذ مالیات",
    "tax_amount": "روند مالیات", "ins_base": "ماخذ بیمه", "ins_employee": "سهم کارگر",
    "ins_employer": "سهم کارفرما", "ins_unemployment": "بیمه بیکاری", "loan": "مبلغ قسط",
    "buy": "خرید از شرکت", "debt_prev": "بدهی از ماه قبل", "debt_misc": "بدهی متفرقه",
    "supp_ins": "بیمه تکمیلی", "fines": "جرائم وخلافی خودروها", "medical": "آزمایش وطب کار",
    "late_base": "ماخد کسر کار - دیر کرد", "late_amount": "مبلغ کسر کار - دیر کرد",
    "store": "کسر انبار", "deductions": "جمع كسورات", "net": "خالص پرداخت",
    "wage_day_prev": "مزدروزانه 1404", "min_wage": "حداقل دستمزد 1405", "housing": "حق مسکن",
    "children": "تعداد اولاد", "child": "حق اولاد", "marriage": "حق تاهل", "food": "وجه بن",
    "ot_real_min": "اضافه کارواقعی دقیقه", "ot_real_hour": "اضافه کار واقعی ساعت",
    "ot_real_amount": "مبلغ اضافه کار واقعی", "excess_fixed": "مازاد ثابت 1405",
    "excess_ot": "اضافه کار مازاد",
    "excess_sum": "جمع مازاد ثابت / اضافه کار ثابت / اضافه کار واقعی ومرخصی",
    "excess_pay": "قابل پرداخت مازاد ثابت / اضافه کار ثابت / اضافه کار واقعی",
    "mission_calc_days": "محاسبه مدت ماموریت", "mission_calc_amount": "محاسبه  مبلغ ماموریت",
    "excess1": "مازاد محاسبه  یک", "ot_rate_hour": "اضافه کار هر ساعت",
    "ot_rate_min": "اضافه کار هردقیقه", "calc_ot_min": "محاسبه اضافه کار دقیقه",
    "calc_ot_hour": "محاسبه اضافه کار ساعت", "calc_ot_min_amount": "مبلغ اضافه کار محاسبه دقیقه",
    "calc_ot_hour_amount": "مبلغ اضافه کار محاسبه ساعت", "excess2": "مازاد محاسبه دو",
    "excess3": "مازاد محاسبه سه", "mission_base": "ماخذ پایانکار وعیدی وپاداش ووجه مرخصی وماموریت",
    "comm_deduct": "کسر مازاد پرداختی فروش", "comm1": "پورسانت یک",
    "leave_prev": "جمع مانده   تا آخر سال  1404 روز", "leave_carry": "مرخصی انتقالی از سال 1404  به روز",
    "leave_year": "مرخصی استحقاق سال 1405 به روز", "leave_total_d": "جمع مر خصی قابل استفاده در سال 1405 به روز",
    "leave_total_h": "مرخصی قابل استفاده سال 1405 به ساعت",
    "leave_total_m": "جمع مرخصی قابل استفاده سال 1405 به دقیقه",
    "used_m": "مصرفی ماه  اعلامی  دقیقه", "used_h": "مصرفی ماه  اعلامی  ساعت", "used_d": "مصرفی ماه  اعلامی  روز",
    "real_m": "مصرفی ماه  واقعی  دقیقه", "real_h": "مصرفی ماه  واقعی  ساعت", "real_d": "مصرفی ماه  واقعی  روز",
    "ytd_m": "مصرفی اعلامی  تا  ماه دقیقه", "ytd_h": "مصرفی اعلامی  تا  ماه ساعت", "ytd_d": "مصرفی اعلامی  تا  ماه روز",
    "ytd_real_m": "مصرفی  تا ماه  دقیقه", "ytd_real_h": "مصرفی تا ماه  ساعت", "ytd_real_d": "مصرفی تا ماه  روز",
    "used_all_m": "استفاده شده کل تا ماه دقیقه", "balance_m": "مانده تا آخر سال  دقیقه",
    "balance_h": "مانده تا آخر سال  ساعت", "balance_d": "مانده تا آخر سال  روز",
}


class Sheet:
    """خواندن یک ماه با نام ستون، نه با حرف ستون."""

    def __init__(self, values, formulas, title):
        self.ws, self.wsf, self.title = values[title], formulas[title], title
        self.index = {}
        for c in range(1, self.ws.max_column + 1):
            head = self.ws.cell(1, c).value
            if head is not None:
                self.index.setdefault(str(head).strip(), c)
        self.last = max(
            r for r in range(2, self.ws.max_row + 1) if self.ws.cell(r, 1).value not in (None, "")
        )

    def col(self, key):
        return self.index.get(COLUMNS[key].strip())

    def get(self, key, row):
        c = self.col(key)
        if c is None:
            return None
        v = self.ws.cell(row, c).value
        return D(str(v)) if isinstance(v, (int, float)) else v

    def num(self, key, row):
        v = self.get(key, row)
        return v if isinstance(v, D) else D("0")

    def has_formula(self, key, row):
        c = self.col(key)
        if c is None:
            return False
        f = self.wsf.cell(row, c).value
        return isinstance(f, str) and f.startswith("=")

    def name(self, row):
        return self.ws.cell(row, 5).value


def check(sheet, row, expected_key, computed, log):
    expected = sheet.get(expected_key, row)
    if expected is None or isinstance(expected, str):
        return 0
    if D(str(expected)) != D(str(computed)):
        log.append((row, sheet.name(row), expected_key, expected, computed))
    return 1


def verify(sheet, prorate_divisor):
    """بدنهٔ فیش، زنجیرهٔ تخصیص مازاد، پورسانت فروش و زنجیرهٔ مرخصی."""
    cells, problems = 0, []
    for r in range(2, sheet.last + 1):
        n = lambda k: sheet.num(k, r)

        # ---------------------------------------------------------- بدنهٔ فیش
        worked = n("days") - (n("absence") + n("unpaid") + n("sick"))
        wage_day = rnd(n("wage_day_prev") * WAGE_FACTOR + WAGE_FIXED)
        sen_day = n("sen_day")
        wage_sen_day = sen_day + wage_day
        salary_month = wage_day * 30            # همیشه ۳۰، نه طول ماه
        sen_month = sen_day * worked
        base_month = sen_month + salary_month   # مبنای اضافه‌کاری و کسر کار
        cum_sen_month = rnd(n("cum_sen_day") * worked)
        salary_pay = wage_day * worked
        sen_pay = sen_day * worked
        sen_diff = rnd(cum_sen_month - sen_pay) if cum_sen_month >= sen_pay else D("0")

        child = n("children") * n("min_wage") * 3
        full = lambda amount: (
            amount if worked in (D(30), D(31)) else rnd(amount / prorate_divisor * worked)
        )
        housing_pay, child_pay = full(n("housing")), full(child)
        marriage_pay, food_pay = full(n("marriage")), full(n("food"))

        ot_amount = rnd(
            base_month / MONTH_HOURS * OT_FACTOR * n("ot_hour")
            + base_month / OT_MINUTES * OT_FACTOR * n("ot_min")
        )
        gross = (
            n("prev_claim") + n("comm3_pay") + n("comm2_pay") + n("comm1_pay") + n("mission_amount")
            + n("eid_pay") + n("end_pay") + n("leave_pay") + ot_amount
            + food_pay + marriage_pay + child_pay + housing_pay + sen_diff + sen_pay + salary_pay
        )
        advance = n("adv_bank") + n("adv_cash")
        tax_base = gross - n("mission_amount")
        tax_amount = rnd(income_tax(tax_base))
        ins_base = gross - (n("mission_amount") + child_pay)
        late_base = salary_month + sen_month + n("housing") + child + n("marriage") + n("food")
        late_amount = rnd(
            late_base / MONTH_HOURS * n("late_h") + late_base / LATE_MINUTES * n("late_m")
        )
        deductions = (
            late_amount + n("medical") + n("fines") + n("supp_ins") + n("debt_misc")
            + n("debt_prev") + n("buy") + n("loan") + rnd(ins_base * D("0.07"))
            + tax_amount + advance + n("store")
        )

        for key, value in [
            ("worked", worked), ("wage_day", wage_day), ("wage_sen_day", wage_sen_day),
            ("salary_month", salary_month), ("sen_month", sen_month),
            ("salary_sen_month", base_month), ("cum_sen_month", cum_sen_month),
            ("salary_pay", salary_pay), ("sen_pay", sen_pay), ("sen_diff", sen_diff),
            ("child", child), ("housing_pay", housing_pay), ("child_pay", child_pay),
            ("marriage_pay", marriage_pay), ("food_pay", food_pay), ("ot_amount", ot_amount),
            ("gross", gross), ("adv_total", advance), ("tax_base", tax_base),
            ("tax_amount", tax_amount), ("ins_base", ins_base),
            ("ins_employee", rnd(ins_base * D("0.07"))),
            ("ins_employer", rnd(ins_base * D("0.20"))),
            ("ins_unemployment", rnd(ins_base * D("0.03"))),
            ("late_base", late_base), ("late_amount", late_amount),
            ("deductions", deductions), ("net", gross - deductions),
        ]:
            cells += check(sheet, r, key, value, problems)

        # ------------------------------------ زنجیرهٔ تخصیص مازاد (گروه پشتیبانی)
        daily = n("mission_base")
        if daily:
            pool = n("ot_real_amount") + n("excess_ot") + n("excess_fixed") + n("leave_pay")
            allocatable = pool - sen_diff
            mission_days = rdown(allocatable / daily)
            mission_amount = mission_days * daily
            rest1 = allocatable - mission_amount
            hour_rate = rnd(base_month / MONTH_HOURS * OT_FACTOR)
            minute_rate = rnd(base_month / OT_MINUTES * OT_FACTOR)
            hours = rdown(rest1 / hour_rate) if hour_rate else D("0")
            hours_amount = rnd(base_month / MONTH_HOURS * OT_FACTOR * hours)
            rest2 = rest1 - hours_amount
            minutes = D("0") if rest2 < minute_rate else rup(rest2 / minute_rate)
            minutes_amount = rnd(base_month / OT_MINUTES * OT_FACTOR * minutes)
            for key, value in [
                ("excess_sum", pool), ("excess_pay", allocatable),
                ("mission_calc_days", mission_days), ("mission_calc_amount", mission_amount),
                ("excess1", rest1), ("ot_rate_hour", hour_rate), ("ot_rate_min", minute_rate),
                ("calc_ot_hour", hours), ("calc_ot_hour_amount", hours_amount), ("excess2", rest2),
                ("calc_ot_min", minutes), ("calc_ot_min_amount", minutes_amount),
                ("excess3", rest2 - minutes_amount),
            ]:
                cells += check(sheet, r, key, value, problems)

        # --------------------------------------- پورسانت و مأموریت (گروه فروش)
        if sheet.has_formula("comm1_pay", r):
            cells += check(
                sheet, r, "comm1_pay",
                (n("comm1") + n("excess_fixed"))
                - (sen_diff + n("comm_reserve") + n("comm_deduct") + n("mission_amount")),
                problems,
            )
            cells += check(sheet, r, "mission_amount", n("mission_days") * wage_sen_day * 2, problems)

        # ------------------------------------------------------ زنجیرهٔ مرخصی
        carry = CARRY_CAP if n("leave_prev") >= CARRY_CAP else rnd(n("leave_prev"))
        total_days = n("leave_year") + carry
        total_minutes = total_days * LEAVE_DAY_MINUTES

        def split(days, hours, minutes):
            """روز/ساعت/دقیقه از روی دقیقه — همان ترتیب ROUNDDOWN فایل."""
            total = days * LEAVE_DAY_MINUTES + hours * 60 + minutes
            d = rdown(total / LEAVE_DAY_MINUTES)
            h = rdown((total - d * LEAVE_DAY_MINUTES) / 60)
            return d, h, rdown(total - (d * LEAVE_DAY_MINUTES + h * 60))

        real_d, real_h, real_m = split(n("used_d"), n("used_h"), n("used_m"))
        ytd_d, ytd_h, ytd_m = split(n("ytd_d"), n("ytd_h"), n("ytd_m"))
        used_all = rdown(ytd_d * LEAVE_DAY_MINUTES + ytd_h * 60 + ytd_m)
        balance_d = rdown((total_minutes - used_all) / LEAVE_DAY_MINUTES)
        balance_h = rdown((total_minutes - used_all - balance_d * LEAVE_DAY_MINUTES) / 60)
        for key, value in [
            ("leave_carry", carry), ("leave_total_d", total_days),
            ("leave_total_h", rdown(total_minutes / 60)),
            ("leave_total_m", rdown(total_minutes - rdown(total_minutes / 60) * 60)),
            ("real_d", real_d), ("real_h", real_h), ("real_m", real_m),
            ("ytd_real_d", ytd_d), ("ytd_real_h", ytd_h), ("ytd_real_m", ytd_m),
            ("used_all_m", used_all), ("balance_d", balance_d), ("balance_h", balance_h),
            ("balance_m", rdown(total_minutes - used_all - (balance_d * LEAVE_DAY_MINUTES + balance_h * 60))),
        ]:
            cells += check(sheet, r, key, value, problems)

    return cells, problems


def main(path):
    values = openpyxl.load_workbook(path, data_only=True)
    formulas = openpyxl.load_workbook(path, data_only=False)
    months = [t for t in values.sheetnames if any(t.startswith(m) for m in (
        "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد",
        "شهریور", "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"))]

    total_cells = total_bad = 0
    for title in months:
        sheet = Sheet(values, formulas, title)
        # فروردین ۱۴۰۵ مزایا را بر ۳۰ تسهیم کرده بود؛ از اردیبهشت بر ۳۱.
        divisor = D("30") if title.startswith("فروردین") else D("31")
        cells, problems = verify(sheet, divisor)
        total_cells += cells
        total_bad += len(problems)
        print(f"{title}: {sheet.last - 1} نفر · {cells} سلول بازسازی شد · مغایرت {len(problems)}")
        for row, name, key, expected, computed in problems:
            print(f"    سطر {row} — {name} · {key}: اکسل {expected} ≠ بازسازی {computed}")

    print(f"\nجمع: {total_cells} سلول · {total_bad} مغایرت")
    return 1 if total_bad else 0


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(2)
    sys.exit(main(sys.argv[1]))
