"""بارگذاری یک ماهِ فایل اکسل شرکت در سامانه، برای مقایسهٔ ریال‌به‌ریال.

    python tools/load_excel_month.py /path/to/1405.xlsx "تیر 1405"

چرا این اسکریپت هست: بهترین آزمونِ یک سامانهٔ حقوق این است که همان ورودی‌هایی
که حسابدار در اکسل وارد می‌کند به سامانه داده شود و خروجی‌ها کنار هم گذاشته
شوند. تفاوت‌ها همان فهرست کارهای باقی‌مانده است.

**فقط ورودی‌ها بارگذاری می‌شوند، نه جواب‌ها.** مزد روزانه، کارکرد، ساعت
اضافه‌کاری، روز مأموریت، پورسانت خام و کسورات وارد می‌شوند؛ حقوق ثابت، بیمه،
مالیات و خالص را خودِ موتور می‌سازد. تنها استثنا آنجاست که سامانه هنوز قاعده‌اش
را ندارد (مابه‌التفاوت پایه سنوات تجمیعی) — آن یکی به‌صورت «مبلغ دستی» وارد
می‌شود، دقیقاً همان کاری که اپراتور امروز می‌کند.

اسکریپت idempotent است: دوباره اجرا کردنش همان نتیجه را می‌دهد.
"""

import hashlib
import os
import re
import sys
from decimal import Decimal as D

import django

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

import jdatetime  # noqa: E402
import openpyxl  # noqa: E402

from apps.attendance.models import Timesheet  # noqa: E402
from apps.employees.models import Dependent, Employee, EmploymentContract  # noqa: E402
from apps.org.models import Company, CostCenter, Department, JobTitle  # noqa: E402
from apps.payroll.company_config import configure  # noqa: E402
from apps.payroll.models import MonthlyParameter, PayrollInput, PayrollPeriod  # noqa: E402
from apps.payroll_config.models import LegalParameter, SalaryComponent  # noqa: E402

MONTH_NAMES = ["فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
               "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"]


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def num(value):
    return D(str(value)) if isinstance(value, (int, float)) else D("0")


def jalali(text):
    """«۱۳۹۵/۰۹/۰۱» → تاریخ میلادی."""
    parts = re.findall(r"\d+", clean(text))
    if len(parts) != 3:
        return None
    y, m, d = (int(x) for x in parts)
    return jdatetime.date(y, m, d).togregorian()


class Sheet:
    # نام ستون بین فایل‌های شرکت فرق می‌کند و هر دو شکل واقعی‌اند. کلید سمت
    # چپ نامی است که این اسکریپت با آن کار می‌کند؛ سمت راست هر چیزی که تا
    # امروز در یک فایل واقعی دیده شده. اولین ستونی که پیدا شود برنده است.
    ALIASES = {
        "تاریخ شروع بکار": ["تاریخ استخدام"],
        "بدهی متفرقه": ["سایر بدهی"],
        "مازاد ثابت 1405": ["مازاد ثابت"],
        "اضافه کار مازاد": ["مازاد پرداختی ماهانه"],
    }

    def __init__(self, path, title, insurance_exempt=False):
        book = openpyxl.load_workbook(path, data_only=True)
        self.ws = book[title]
        self.title = title
        # همهٔ پرسنل این شیت بدون بیمه‌اند. در فایل تیر ۱۴۰۵ این تنها تفاوت
        # شیت دوم است: ستون «ماخذ بیمه» صفرِ سخت است و بقیهٔ فرمول‌ها یکی‌اند.
        self.insurance_exempt = insurance_exempt

        # سطر عنوان همیشه سطر ۱ نیست؛ از روی ستون نام پیدا می‌شود.
        self.header_row = self._find_header()
        self.index = {}
        for c in range(1, self.ws.max_column + 1):
            head = clean(self.ws.cell(self.header_row, c).value)
            if head:
                self.index.setdefault(head, c)
        self.rows = [
            r for r in range(self.header_row + 1, self.ws.max_row + 1)
            if self.ws.cell(r, self.index["نام ونام خانوادگي"]).value
        ]

    def _find_header(self):
        for r in range(1, 6):
            for c in range(1, self.ws.max_column + 1):
                if clean(self.ws.cell(r, c).value) == "نام ونام خانوادگي":
                    return r
        raise SystemExit("ستون «نام ونام خانوادگي» پیدا نشد — آیا شیت درست است؟")

    def column(self, key):
        """شمارهٔ ستون، با در نظر گرفتن نام‌های جایگزین."""
        if key in self.index:
            return self.index[key]
        for alias in self.ALIASES.get(key, ()):
            if alias in self.index:
                return self.index[alias]
        return None

    def get(self, key, row):
        c = self.column(key)
        return self.ws.cell(row, c).value if c else None

    def n(self, key, row):
        return num(self.get(key, row))

    def s(self, key, row):
        return clean(self.get(key, row))


# ---------------------------------------------------------------- ساختار سازمانی

def ensure_org(company, sheet):
    """واحد و پست سازمانی را از خود فایل می‌سازد.

    مرکز هزینه در فایل تیر خالی است، پس واحد هم‌نام مرکز هزینه گرفته می‌شود —
    گزارش‌های سازمانی بدون مرکز هزینه معنا ندارند.
    """
    units = sorted({sheet.s("واحد", r) for r in sheet.rows if sheet.s("واحد", r)})
    titles = sorted({sheet.s("شغل", r) for r in sheet.rows if sheet.s("شغل", r)})

    centers, departments = {}, {}
    for name in units:
        center, _ = CostCenter.objects.get_or_create(company=company, name=name)
        centers[name] = center
        dept, _ = Department.objects.get_or_create(
            company=company, name=name, defaults={"cost_center": center}
        )
        if dept.cost_center_id != center.pk:
            dept.cost_center = center
            dept.save(update_fields=["cost_center"])
        departments[name] = dept

    jobs = {}
    for name in titles:
        jobs[name], _ = JobTitle.objects.get_or_create(company=company, name=name)
    return departments, centers, jobs


# ---------------------------------------------------------------- پارامترهای سال

def ensure_params(sheet, fiscal_year):
    """مبالغ سال را از خود فایل برمی‌دارد، نه از حدس."""
    row = sheet.rows[0]
    params = LegalParameter.objects.filter(fiscal_year=fiscal_year).first()
    params.min_daily_wage = sheet.n("حداقل دستمزد 1405", row)
    params.housing_allowance = sheet.n("حق مسکن", row)
    params.food_allowance = sheet.n("وجه بن", row)
    params.seniority_daily = sheet.n("پایه سنوات روزانه", row) or D("166667")

    # حق اولاد و حق تأهل باید از سطری خوانده شوند که واقعاً دارندشان
    for r in sheet.rows:
        if sheet.n("تعداد اولاد", r):
            params.child_allowance = sheet.n("حق اولاد", r) / sheet.n("تعداد اولاد", r)
            break
    for r in sheet.rows:
        if sheet.n("حق تاهل", r):
            params.marriage_allowance = sheet.n("حق تاهل", r)
            break
    params.save()
    return params


# ---------------------------------------------------------------- هویت پرسنل

# کلید پایدار هر سطر: (عنوان شیت، شماره سطر) → کد پرسنلی که سامانه به کار می‌برد
KEY_MAP = {}
# کدهایی که در فایل بین دو نفر مشترک‌اند، و سطرهایی که اصلاً کد ندارند
CODE_PROBLEMS = {"duplicate": {}, "missing": []}
# کسانی که فایل به آن‌ها پایه سنوات می‌دهد ولی تاریخ استخدامشان یک سال نشده
PRIOR_SERVICE = []


def _name_key(full_name: str) -> str:
    """کد جایگزین برای کسی که در فایل کد پرسنلی ندارد.

    از نام ساخته می‌شود و نه از شمارهٔ سطر: سطرِ همان آدم ماه بعد جابه‌جا
    می‌شود و یک پرسنل تکراری ساخته می‌شود. هش کوتاهِ نام هم پایدار است و هم
    در ۲۰ کاراکتر جا می‌شود (نام‌های فارسی جا نمی‌شوند).

    این یک **راه‌حل موقت** است؛ عدد واقعی باید از شرکت گرفته شود.
    """
    digest = hashlib.sha1(full_name.encode("utf-8")).hexdigest()[:8]
    return f"N-{digest}"


def build_key_map(sheets):
    """کد پرسنلیِ قابل اتکا برای هر سطر — و گزارش آنچه در فایل خراب است.

    فایل تیر ۱۴۰۵ دو مشکل هویتی دارد که هر دو به فیشِ اشتباه ختم می‌شوند:

      ۱) ۱۶ سطر شماره پرسنلی ندارند (خانه‌شان صفر است)
      ۲) سه کد بین **دو نفر متفاوت** مشترک‌اند

    اگر کد را همان‌طور که هست بگیریم، آن دو نفر یک پرسنل می‌شوند و فیش دومی
    روی اولی می‌نشیند. پس کد فقط وقتی پذیرفته می‌شود که در کل فایل یکتا باشد؛
    وگرنه نام — که در هر ۶۷ سطر یکتاست — مبنا می‌شود.
    """
    KEY_MAP.clear()
    PRIOR_SERVICE.clear()
    CODE_PROBLEMS["duplicate"].clear()
    CODE_PROBLEMS["missing"].clear()

    seen = {}
    for sheet in sheets:
        for row in sheet.rows:
            code = sheet.s("شماره پرسنلی", row)
            code = "" if code in {"", "0"} else code
            name = sheet.s("نام ونام خانوادگي", row)
            seen.setdefault(code, []).append((sheet, row, name))

    for code, entries in seen.items():
        shared = code and len(entries) > 1
        if shared:
            CODE_PROBLEMS["duplicate"][code] = [name for _, _, name in entries]
        for sheet, row, name in entries:
            if not code:
                CODE_PROBLEMS["missing"].append(name)
            KEY_MAP[(sheet.title, row)] = code if (code and not shared) else _name_key(name)
    return KEY_MAP


def personnel_code(sheet, row) -> str:
    return KEY_MAP[(sheet.title, row)]


def placeholder_national_id(code: str) -> str:
    """کد ملیِ جایگزین — یکتا، ده رقمی، و آشکارا غیرواقعی.

    فایل کد ملی ندارد ولی این فیلد در مدل یکتاست. با «۹» شروع می‌شود تا با
    کد ملی واقعی اشتباه گرفته نشود، و از خودِ کد پرسنلی ساخته می‌شود تا هر بار
    همان عدد دربیاید.
    """
    digest = hashlib.sha1(code.encode("utf-8")).hexdigest()
    return "9" + str(int(digest[:12], 16))[:9].rjust(9, "0")


# ---------------------------------------------------------------- پرسنل

def load_people(company, sheet, departments, centers, jobs, period):
    created = updated = 0
    for r in sheet.rows:
        full = sheet.s("نام ونام خانوادگي", r)
        first, _, last = full.partition(" ")
        # چند سطر فایل شماره پرسنلی ندارند. شمارهٔ سطر به‌عنوان کد جایگزین
        # **غلط** است: سطر همان آدم در ماه بعد جابه‌جا می‌شود و یک پرسنل تکراری
        # ساخته می‌شود. نام، پایدارتر است.
        code = personnel_code(sheet, r)
        hire = jalali(sheet.get("تاریخ شروع بکار", r))
        gender = "F" if "خانم" in sheet.s("آقا / خانم", r) else "M"
        married = "متاهل" in sheet.s("وضعیت تاهل", r)

        # hire_date باید در defaults باشد: get_or_create همان لحظه INSERT می‌زند
        # و این فیلد روی مدل اجباری است.
        employee, is_new = Employee.objects.get_or_create(
            company=company, personnel_code=code,
            defaults={
                "first_name": first,
                "last_name": last.strip() or first,
                "hire_date": hire,
                # فایل کد ملی ندارد و این فیلد در مدل یکتاست. کد پرسنلیِ
                # صفرپرشده گذاشته می‌شود: یکتا هست و آشکارا کد ملی واقعی نیست،
                # پس کسی اشتباهی جایی از آن استفاده نمی‌کند.
                "national_id": placeholder_national_id(code),
            },
        )
        employee.first_name = first
        employee.last_name = last.strip() or first
        employee.gender = gender
        employee.marital_status = "MARRIED" if married else "SINGLE"
        employee.hire_date = hire
        employee.status = Employee.Status.ACTIVE
        # شیت «بدون بیمه» گروه محاسباتی جدایی نیست؛ فقط ماخذ بیمه‌اش صفر است.
        # همان تیکِ روی خودِ پرسنل این را می‌گوید، بدون هیچ قلم یا قاعدهٔ تازه.
        employee.is_insurance_exempt = sheet.insurance_exempt
        if sheet.insurance_exempt and not employee.insurance_exempt_reason:
            employee.insurance_exempt_reason = f"شیت «{sheet.title}»"

        # سابقهٔ قبلی که در فایل ستون ندارد ولی فایل خودش شهادت می‌دهد هست:
        # اگر «سنوات قابل پرداخت» پر باشد در حالی که تاریخ استخدام هنوز یک سال
        # نشده، یعنی این نفر سابقهٔ همان کارگاه را دارد و جایی ثبت نشده.
        # کمترین مقداری که او را به یک سال می‌رساند گذاشته می‌شود — عددی که
        # فایل را می‌خواند، ولی **باید با پروندهٔ پرسنلی تأیید شود**.
        if sheet.n("سنوات قابل پرداخت", r) > 0:
            months = employee.seniority_months_at(period.end_date)
            if months < 12:
                top_up = 12 - months
                employee.prior_service_months = (
                    employee.prior_service_months or 0
                ) + top_up
                PRIOR_SERVICE.append((employee.full_name, top_up))
        employee.save()
        created += is_new
        updated += not is_new

        # فرزندان: فقط تعدادشان در فایل هست، پس همان تعداد رکورد ساخته می‌شود.
        #
        # تاریخ شروع تکفل، **شروع همین دوره** است نه تاریخ استخدام: تعداد فرزند
        # بین ماه‌ها فرق می‌کند و موتور حق اولاد را با تاریخ حساب می‌کند. اگر
        # تاریخ استخدام بگذاریم، فرزندی که در اردیبهشت اضافه شده در فروردین هم
        # حق اولاد می‌گیرد.
        want = int(sheet.n("تعداد اولاد", r))
        have = employee.dependents.filter(relation=Dependent.Relation.CHILD).count()
        for i in range(have, want):
            Dependent.objects.create(
                employee=employee, first_name=f"فرزند {i + 1}", last_name=employee.last_name,
                relation=Dependent.Relation.CHILD, start_date=period.start_date,
                child_allowance_eligible=True,
            )
        if have > want:
            for extra in employee.dependents.filter(
                relation=Dependent.Relation.CHILD
            ).order_by("-pk")[: have - want]:
                extra.delete()

        unit = sheet.s("واحد", r)
        job = sheet.s("شغل", r)
        years = max(0, (period.end_date - hire).days // 365) if hire else 0
        contract, _ = EmploymentContract.objects.get_or_create(
            employee=employee,
            defaults={
                "effective_from": hire,
                "status": EmploymentContract.Status.ACTIVE,
                "department": departments[unit],
                "cost_center": centers[unit],
                "job_title": jobs[job],
            },
        )
        contract.contract_type = EmploymentContract.ContractType.PERMANENT
        contract.department = departments[unit]
        contract.cost_center = centers[unit]
        contract.job_title = jobs[job]
        contract.effective_from = hire
        contract.daily_wage = sheet.n("مزدروزانه", r)
        contract.status = EmploymentContract.Status.ACTIVE
        contract.save()

        _bank_account(sheet, r, employee)
    return created, updated


def _bank_account(sheet, row, employee):
    """حساب بانکی — چون روی فیش چاپ می‌شود و فایل پرداخت بانک از آن ساخته می‌شود.

    شمارهٔ حساب در فایل با نقطه یا اسلش نوشته شده (۱۰.۸۵۶۹۷۳۵.۱ و ۱۰/۸۳۴۷۷۱۹/۱)
    و هر دو یک چیزند؛ به یک شکل درمی‌آیند وگرنه فایل بانک دو ردیف متفاوت
    می‌بیند. «۰» یعنی خالی، نه شمارهٔ حساب صفر.
    """
    from apps.employees.models import BankAccount

    number = clean(sheet.get("شماره حساب", row)).replace("/", ".")
    if number in {"", "0"}:
        return
    card = clean(sheet.get("شماره کارت", row))
    bank = clean(sheet.get("بانك عامل", row)) or BankAccount.DEFAULT_BANK

    account, _ = BankAccount.objects.get_or_create(
        employee=employee, account_number=number,
        defaults={"bank_name": bank, "is_default": True},
    )
    account.bank_name = bank
    account.card_number = "" if card in {"", "0"} else card
    account.is_default = True
    account.is_active = True
    account.save()
    # هر پرسنل فقط یک حساب پیش‌فرض
    employee.bank_accounts.exclude(pk=account.pk).update(is_default=False)


# ---------------------------------------------------------------- کارکرد

def load_timesheets(sheet, period, params):
    day_minutes = params.leave_day_minutes or 440
    for r in sheet.rows:
        employee = _employee(sheet, r, period)
        contract = employee.contracts.filter(status="ACTIVE").first()
        timesheet, _ = Timesheet.objects.get_or_create(
            period=period, employee=employee, defaults={"contract": contract}
        )
        timesheet.contract = contract
        # `work_days` کارکرد **اولیه** است و موتور خودش غیبت و بیماری و مرخصی
        # بدون حقوق را از آن کم می‌کند. ستون «کارکرد واقعی ماهانه» در فایل
        # همان کسرشده است (O = K − L − M − N)، پس اگر آن را بگذاریم و سه ستون
        # را هم جدا بدهیم، روزهای بی‌حقوق دو بار کم می‌شوند.
        timesheet.work_days = sheet.n("روز های ماه", r) or sheet.n("کارکرد واقعی ماهانه", r)
        timesheet.absence_days = sheet.n("غیبت", r)
        timesheet.unpaid_leave_days = sheet.n("بدون حقوق", r)
        timesheet.sick_leave_days = sheet.n("بیماری", r)
        # اضافه‌کاری و روز مأموریت: کدام ستون **ورودی** است، به گروه بستگی دارد.
        #
        # فروش    → «اضافه کاری عادی» و «مدت ماموریت» خودشان ورودی‌اند.
        # پشتیبانی → این دو **خروجی** زنجیرهٔ مازادند، نه ورودی. ورودی واقعی
        #            «اضافه کار واقعی» است و روز مأموریت را سامانه می‌سازد.
        #
        # اگر خروجی را به‌عنوان ورودی بار کنیم، اعداد می‌خوانند ولی زنجیره
        # آزموده نمی‌شود — یعنی بک‌تست خودش را تأیید می‌کند نه سامانه را.
        if is_sales(sheet, r):
            timesheet.mission_days = sheet.n("مدت ماموریت", r)
            timesheet.overtime_minutes = int(
                sheet.n("اضافه کاری عادی ساعت", r) * 60
                + sheet.n("اضافه کاری عادی دقیقه", r)
            )
        else:
            timesheet.mission_days = 0
            timesheet.overtime_minutes = int(
                sheet.n("اضافه کار واقعی ساعت", r) * 60
                + sheet.n("اضافه کارواقعی دقیقه", r)
            )
        # مرخصی استحقاقیِ همین ماه — روز و ساعت و دقیقه، به دقیقه تبدیل‌شده.
        # فایل «مصرفی ماه اعلامی» را ورودی می‌داند و «واقعی» را از رویش می‌سازد،
        # پس اعلامی ورودی است.
        timesheet.leave_minutes = int(
            sheet.n("مصرفی ماه اعلامی روز", r) * day_minutes
            + sheet.n("مصرفی ماه اعلامی ساعت", r) * 60
            + sheet.n("مصرفی ماه اعلامی دقیقه", r)
        )
        timesheet.status = Timesheet.Status.APPROVED
        timesheet.save()

        _leave_entitlement(
            sheet, r, employee, period, day_minutes, timesheet.leave_minutes
        )


def _leave_entitlement(sheet, row, employee, period, day_minutes, month_minutes):
    """سهمیهٔ مرخصی سال، به‌علاوهٔ مانده‌گیری افتتاحیه.

    «مصرفی تا ماه» و «مانده» عمداً ذخیره نمی‌شوند: سامانه هر دو را از جمع
    کارکرد ماه‌های همان سال می‌سازد، تا اصلاح کارکرد خرداد خودبه‌خود مانده را
    درست کند.

    ولی سامانه از **وسط سال** عملیاتی شده و ماه‌های قبلش کارکردی در سامانه
    ندارند. ستون «مصرفی اعلامی تا ماه» در فایل، مصرفیِ از اول سال **شاملِ
    همین ماه** است؛ پس سهمِ ماه‌های قبل همان عدد منهای مرخصی همین ماه است، و
    مرزش ماهِ قبل.

    وقتی روزی فروردین تا خرداد واقعاً وارد سامانه شدند، کافی است مرز صفر شود
    تا همه‌چیز دوباره از کارکرد واقعی ساخته شود.
    """
    from apps.attendance.models import LeaveEntitlement

    carried = int(sheet.n("مرخصی انتقالی از سال 1404 به روز", row) * day_minutes)
    annual = int(sheet.n("مرخصی استحقاق سال 1405 به روز", row) * day_minutes)

    declared_ytd = int(
        sheet.n("مصرفی اعلامی تا ماه روز", row) * day_minutes
        + sheet.n("مصرفی اعلامی تا ماه ساعت", row) * 60
        + sheet.n("مصرفی اعلامی تا ماه دقیقه", row)
    )
    opening = max(declared_ytd - int(month_minutes or 0), 0)
    through = period.month - 1 if opening else 0

    if not carried and not annual and not opening:
        return
    LeaveEntitlement.objects.update_or_create(
        employee=employee, fiscal_year=period.fiscal_year,
        defaults={
            "carried_over_minutes": carried,
            "annual_minutes": annual,
            "opening_used_minutes": opening,
            "opening_through_month": through,
            "note": f"از فایل «{sheet.title}»",
        },
    )


# ---------------------------------------------------------------- وام و اقساط

def load_loans(sheet, period):
    """قسطِ همین ماه را به‌صورت یک وام واقعی ثبت می‌کند، نه کسور دستی.

    فایل فقط «مبلغ قسط» این ماه را دارد (و گاهی تعداد کل و شمارهٔ قسط). همان
    را وام یک‌قسطیِ همین دوره می‌سازیم تا قاعدهٔ `loan_installment` موتور
    واقعاً آزموده شود؛ اگر به‌صورت کسور دستی وارد می‌شد، عدد می‌خواند ولی
    هیچ‌چیزی از سامانه تست نشده بود.

    اقساط هر بار به «سررسید» برمی‌گردند: محاسبهٔ قبلی آن‌ها را «کسر شده» کرده
    و اجرای دوم دیگر نمی‌دیدشان — یعنی بار دوم عدد وام از فیش غایب می‌شد.
    """
    from apps.loans.models import Loan, LoanInstallment

    count = 0
    for r in sheet.rows:
        amount = sheet.n("مبلغ قسط", r)
        if not amount:
            continue
        employee = _employee(sheet, r, period)
        loan, _ = Loan.objects.get_or_create(
            employee=employee, first_year=period.year, first_month=period.month,
            defaults={
                "loan_type": Loan.LoanType.LOAN,
                "title": f"قسط {period.title} — از فایل اکسل",
                "principal": amount,
                "total_installments": 1,
                "installment_amount": amount,
            },
        )
        loan.principal = amount
        loan.installment_amount = amount
        loan.status = Loan.Status.ACTIVE
        loan.save()
        installment, _ = LoanInstallment.objects.get_or_create(
            loan=loan, number=1,
            defaults={"due_year": period.year, "due_month": period.month, "amount": amount},
        )
        installment.amount = amount
        installment.due_year, installment.due_month = period.year, period.month
        installment.status = LoanInstallment.Status.DUE
        installment.payslip_line = None
        installment.save()
        count += 1
    return count


# ---------------------------------------------------------------- مبالغ دستی

# ستون اکسل → کد قلم حقوقی. فقط چیزهایی که سامانه قاعده‌ای برایشان ندارد.
MANUAL = {
    "مابه التفاوت پایه سنوات تجمیعی قابل پرداخت": "DIFF",
    # «مازاد ثابت» روی فیش سطر ندارد و فقط داخل فرمول پورسانت است، پس با
    # پورسانت خام یک‌جا وارد می‌شود (ستون ترکیبی پایین‌تر ساخته می‌شود).
    "پورسانت یک": "COMM_1",
    "ذخیره پورسانت فروش": "COMM_RESERVE",
    "پورسانت دو": "COMM_2",
    "پورسانت سه": "COMM_3",
    "طلب ماه قبل": "PREV_CLAIM",
    "وجه مرخصی": "LEAVE_PAY",
    # هزینه تلفن **مشمول بیمه** است؛ از فیش تیر ۱۴۰۵ آقای سعادتی تأیید شد
    # (بدون آن بیمه ۱۵٬۷۴۳٬۵۷۷ می‌شد نه ۱۵٬۸۱۳٬۵۷۷ که در فیش چاپ شده).
    "هزینه تلفن": "PHONE",
    # فقط راننده استجاری: عیدی و پایانکار ماهانه پرداخت می‌شوند نه سالانه
    "پایانکار قابل پرداخت": "SEVERANCE",
    "عیدی وپاداش قابل پرداخت": "EID",
    "علی الحساب واریزی": "ADVANCE",
    "خرید از شرکت": "GOODS",
    "بدهی از ماه قبل": "PREV_DEBT",
    "بدهی متفرقه": "OTHER_DEBT",
    "بیمه تکمیلی": "SUPP_INS",
    "جرائم وخلافی خودروها": "VEHICLE_FINE",
    "آزمایش وطب کار": "MEDICAL",
    "مبلغ کسر کار - دیر کرد": "LATE",
    "کسر انبار": "STORE",
}

# ورودی‌های استخر مازاد — فقط برای گروه پشتیبانی. برای فروش، «مازاد ثابت»
# داخل پورسانت خام می‌رود (پایین‌تر) و اصلاً استخری در کار نیست.
SURPLUS = {
    "مازاد ثابت 1405": "SURPLUS_FIXED",
    "اضافه کار مازاد": "SURPLUS_OT",
}


def is_sales(sheet, row) -> bool:
    return "فروش" in sheet.s("واحد", row)


def _employee(sheet, row, period):
    return Employee.objects.get(
        company=period.company, personnel_code=personnel_code(sheet, row)
    )


def load_manual(sheet, period):
    codes = {c.code: c for c in SalaryComponent.objects.filter(company=period.company)}
    written = skipped = 0
    for r in sheet.rows:
        employee = _employee(sheet, r, period)
        sales = is_sales(sheet, r)
        columns = dict(MANUAL) if sales else {**MANUAL, **SURPLUS}
        for column, code in columns.items():
            component = codes.get(code)
            if component is None:
                skipped += 1
                continue
            value = sheet.n(column, r)
            if code == "COMM_1" and sales:
                # پورسانت خام + مازاد ثابت — همان چیزی که فرمول اکسل جمع می‌کند
                value += sheet.n("مازاد ثابت 1405", r)
            if not value:
                PayrollInput.objects.filter(
                    period=period, employee=employee, component=component
                ).delete()
                continue
            PayrollInput.objects.update_or_create(
                period=period, employee=employee, component=component,
                defaults={"amount": value},
            )
            written += 1
    return written, skipped


def resolve_sheets(path, title):
    """شیت‌های این ماه را پیدا می‌کند — چون یک ماه می‌تواند دو شیت داشته باشد.

    فایل تیر ۱۴۰۵ شرکت ماه را به «با بیمه» و «بدون بیمه» تقسیم کرده. هر دو
    یک دوره‌اند و یک فیش می‌دهند؛ تنها تفاوتشان این است که در شیت دوم ماخذ
    بیمه صفر است. پس هر دو بار می‌شوند و پرسنل شیت دوم تیکِ «بدون بیمه»
    می‌خورند.
    """
    book = openpyxl.load_workbook(path, read_only=True)
    names = list(book.sheetnames)
    book.close()

    if title in names:
        return [(title, "بدون بیمه" in title)]

    month = title.split()[0]
    found = [name for name in names if month in name]
    if not found:
        raise SystemExit(
            f"شیتی به نام «{title}» نیست و هیچ شیتی هم «{month}» ندارد. "
            f"شیت‌های موجود: {'، '.join(names)}"
        )
    return [(name, "بدون بیمه" in name) for name in found]


def main(path, title):
    company = Company.objects.first()
    month = MONTH_NAMES.index(title.split()[0]) + 1
    year = int(title.split()[1])

    sheets = [Sheet(path, name, exempt) for name, exempt in resolve_sheets(path, title)]
    primary = next((s for s in sheets if not s.insurance_exempt), sheets[0])
    build_key_map(sheets)

    fiscal = company.fiscal_years.get(year=year)
    period = PayrollPeriod.objects.filter(company=company, year=year, month=month).first()
    if period is None:
        from apps.payroll.utils import jalali_month_range

        start, end = jalali_month_range(year, month)
        period = PayrollPeriod.objects.create(
            company=company, fiscal_year=fiscal, year=year, month=month,
            start_date=start, end_date=end, status=PayrollPeriod.Status.DRAFT,
        )
    period.status = PayrollPeriod.Status.DRAFT
    period.save()

    params = ensure_params(primary, fiscal)

    created = updated = 0
    for sheet in sheets:
        departments, centers, jobs = ensure_org(company, sheet)
        made, touched = load_people(company, sheet, departments, centers, jobs, period)
        created += made
        updated += touched

    # پیکربندی **بعد از** پرسنل: دامنهٔ زنجیرهٔ مازاد از پست‌های واقعیِ
    # قراردادها ساخته می‌شود، و آن‌ها تا این لحظه وجود نداشتند.
    changes = configure(company)

    written = skipped = loans = 0
    for sheet in sheets:
        load_timesheets(sheet, period, params)
        loans += load_loans(sheet, period)
        made, missed = load_manual(sheet, period)
        written += made
        skipped += missed

    print(f"دوره      : {period.title}  ({period.month_days} روز)")
    print("شیت‌ها     : " + "، ".join(
        f"{s.title}{' (بدون بیمه)' if s.insurance_exempt else ''}" for s in sheets
    ))
    print(f"پرسنل     : {created} تازه · {updated} به‌روز")
    print(f"کارکرد    : {Timesheet.objects.filter(period=period).count()} سطر")
    print(f"وام       : {loans} قسط سررسیدشده")
    print(f"مبالغ دستی: {written} مقدار" + (f" · {skipped} قلمِ نبود" if skipped else ""))
    print(f"پارامترها : حداقل {params.min_daily_wage:,.0f} · مسکن {params.housing_allowance:,.0f}"
          f" · بن {params.food_allowance:,.0f} · اولاد {params.child_allowance:,.0f}"
          f" · تأهل {params.marriage_allowance:,.0f} · سنوات روزانه {params.seniority_daily:,.0f}")
    if changes:
        print("پیکربندی  :")
        for line in changes:
            print(f"   - {line}")

    if PRIOR_SERVICE:
        print()
        print("سابقهٔ قبلی که از فایل استنباط شد (باید با پرونده تأیید شود):")
        for full_name, months in PRIOR_SERVICE:
            print(f"  ! {full_name}: {months} ماه — فایل پایه سنوات می‌دهد "
                  f"ولی تاریخ استخدامش یک سال نشده")

    missing = CODE_PROBLEMS["missing"]
    duplicate = CODE_PROBLEMS["duplicate"]
    if missing or duplicate:
        print()
        print("هویت پرسنل — این دو مورد در خودِ فایل‌اند و باید اصلاح شوند:")
    if duplicate:
        print(f"  ! {len(duplicate)} کد پرسنلی بین دو نفر مشترک است:")
        for code, names in duplicate.items():
            print(f"      {code} → " + " و ".join(names))
        print("      هر دو با کدِ ساخته‌شده از نام وارد شدند، وگرنه فیش یکی روی دیگری می‌نشست.")
    if missing:
        print(f"  ! {len(missing)} نفر شماره پرسنلی ندارند: " + "، ".join(missing[:5])
              + (" و …" if len(missing) > 5 else ""))


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        raise SystemExit(1)
    main(sys.argv[1], sys.argv[2])
