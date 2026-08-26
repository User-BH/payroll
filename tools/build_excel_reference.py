"""ساخت نگاشت «قلم حقوقی → فرمول اکسل» برای صفحهٔ فرمول‌های سامانه.

    python tools/build_excel_reference.py 1405.xlsx "تیر 1405" > apps/payroll/excel_reference.py

سمتِ اکسلِ آن صفحه، عمداً یک **تصویر ثابت** است: فایل شرکت روی سرور نیست و
نباید باشد. هر وقت فایل عوض شد، همین دستور دوباره اجرا می‌شود.

سمتِ سامانه ولی زنده است — از خودِ موتور خوانده می‌شود، نه از این فایل.
"""

import os
import re
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import openpyxl  # noqa: E402


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalise(formula, row):
    text = re.sub(r"\s+", " ", str(formula)).strip()
    return re.sub(rf"(?<=[A-Z]){row}\b", "#", text)


def main(path, title):
    from tools.compare_excel_platform import LINES

    book = openpyxl.load_workbook(path, data_only=False)
    sheet = next((n for n in book.sheetnames if title.split()[0] in n), None)
    if sheet is None:
        raise SystemExit(f"شیتی برای «{title}» پیدا نشد")
    ws = book[sheet]

    head = 1
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            if clean(ws.cell(r, c).value) == "نام ونام خانوادگي":
                head = r
    index = {}
    for c in range(1, ws.max_column + 1):
        name = clean(ws.cell(head, c).value)
        if name:
            index.setdefault(name, c)
    rows = [r for r in range(head + 1, ws.max_row + 1) if ws.cell(r, 5).value]

    def dominant(column):
        """(فرمول غالب، چند سطر، کل سطرها) — یا عددِ دستی بودن.

        ستونی که فرمول ندارد «بی‌ستون» نیست: در خودِ اکسل هم عددِ دستی است،
        و همین را باید گفت. فرقشان مهم است — یکی یعنی «تطبیق ندادیم»، دیگری
        یعنی «هر دو طرف دستی‌اند».
        """
        col = index.get(column)
        if not col:
            return None
        shapes = {}
        constants = 0
        for r in rows:
            value = ws.cell(r, col).value
            if isinstance(value, str) and value.startswith("="):
                key = normalise(value, r)
                shapes[key] = shapes.get(key, 0) + 1
            elif value not in (None, ""):
                constants += 1
        if not shapes:
            return "(عددِ دستی — در اکسل هم فرمول ندارد)", constants, len(rows)
        best = max(shapes.items(), key=lambda x: x[1])
        return best[0], best[1], len(rows)

    print('"""فرمول‌های فایل اکسل شرکت — تصویرِ ثابت، برای مقایسه با موتور.')
    print()
    print(f"از «{os.path.basename(path)}» شیت «{sheet}» ساخته شده در {date.today()}.")
    print()
    print("این فایل با دست ویرایش نمی‌شود:")
    print()
    print("    python tools/build_excel_reference.py <فایل> <ماه> > apps/payroll/excel_reference.py")
    print()
    print("سمتِ سامانه در صفحهٔ فرمول‌ها از خودِ موتور خوانده می‌شود، نه از اینجا.")
    print('"""')
    print()
    print(f'SOURCE = "{os.path.basename(path)} — {sheet}"')
    print()
    print("# کد قلم → (ستون اکسل، فرمول غالب، چند سطر از چند سطر)")
    print("EXCEL_FORMULAS = {")
    for codes, column in LINES:
        keys = codes if isinstance(codes, tuple) else (codes,)
        found = dominant(column)
        if not found:
            continue
        formula, count, total = found
        for code in keys:
            print(f'    "{code}": (')
            print(f'        "{column}",')
            print(f'        "{formula}",')
            print(f"        ({count}, {total}),")
            print("    ),")
    print("}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        raise SystemExit(1)
    main(sys.argv[1], sys.argv[2])
