"""ورود گروهی پرسنل از اکسل.

تایپ دستی ۱۴۶ نفر در چهارده ستون شدنی نیست؛ برای عملیاتی شدن، ورود از اکسل
لازم است. هر سطر یا کامل و بی‌خطا وارد می‌شود یا با ذکر دلیل رد می‌شود، و کل
عملیات داخل یک تراکنش است تا فایل نیمه‌وارد نشود.
"""

import io
from datetime import date
from decimal import Decimal

from django.db import transaction

from apps.employees.models import (
    BankAccount,
    Employee,
    EmploymentContract,
    is_valid_national_id,
)
from apps.org.models import CostCenter, Department, JobTitle
from apps.payroll.utils import jalali_to_gregorian, parse_decimal

COLUMNS = [
    ("personnel_code", "کد پرسنلی", True),
    ("first_name", "نام", True),
    ("last_name", "نام خانوادگی", True),
    ("father_name", "نام پدر", False),
    ("national_id", "کد ملی", True),
    ("birth_date", "تاریخ تولد (۱۳۷۰/۰۵/۱۲)", False),
    ("gender", "جنسیت (مرد/زن)", False),
    ("marital_status", "تأهل (مجرد/متأهل)", False),
    ("insurance_number", "شماره بیمه", False),
    ("mobile", "موبایل", False),
    ("hire_date", "تاریخ استخدام (۱۴۰۲/۰۳/۰۱)", True),
    ("department", "واحد سازمانی", True),
    ("cost_center", "مرکز هزینه", True),
    ("job_title", "پست سازمانی", True),
    ("base_salary", "حقوق ثابت ماهانه (ریال)", True),
    ("seniority_years", "سابقه (سال)", False),
    ("bank_name", "بانک", False),
    ("iban", "شماره شبا", False),
]

GENDERS = {"مرد": "M", "زن": "F", "M": "M", "F": "F"}
MARITAL = {"مجرد": "SINGLE", "متأهل": "MARRIED", "متاهل": "MARRIED"}


def build_template() -> bytes:
    """فایل نمونه با سرستون‌ها و یک سطر راهنما."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "پرسنل"
    sheet.sheet_view.rightToLeft = True

    headers = [label for _, label, _ in COLUMNS]
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="F3EEE4")
    for index, (_, _, required) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True, color="A81812" if required else "1B1815")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        sheet.column_dimensions[cell.column_letter].width = max(len(cell.value) + 4, 14)
    sheet.row_dimensions[1].height = 34

    sheet.append([
        "1001", "علی", "محمدی", "حسن", "0012345678", "1370/05/12", "مرد", "متأهل",
        "1234567890", "09121234567", "1402/03/01", "فروش", "فروش", "بازاریاب",
        "180000000", "3", "ملت", "IR120120000000123456789012",
    ])
    for cell in sheet[2]:
        cell.font = Font(italic=True, color="8A8175")

    sheet.append([])
    sheet.append(["سرستون‌های قرمز اجباری‌اند. سطر دوم فقط نمونه است و باید حذف شود."])
    sheet.cell(row=4, column=1).font = Font(bold=True, color="A81812")

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _parse_jalali(value):
    """«۱۴۰۲/۰۳/۰۱» یا تاریخ اکسل → date میلادی."""
    if value in (None, ""):
        return None
    if isinstance(value, date):
        return value
    text = str(value).strip().translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789"))
    parts = [p for p in text.replace("-", "/").split("/") if p]
    if len(parts) != 3:
        raise ValueError(f"تاریخ «{value}» قابل خواندن نیست")
    year, month, day = (int(p) for p in parts)
    if year < 1500:  # جلالی
        return jalali_to_gregorian(year, month, day)
    return date(year, month, day)


@transaction.atomic
def import_employees(file_obj, company, create_contracts=True) -> dict:
    """خواندن فایل و ثبت پرسنل. یا همه یا هیچ."""
    from openpyxl import load_workbook

    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook.active

    departments = {d.name.strip(): d for d in Department.objects.filter(company=company)}
    cost_centers = {c.name.strip(): c for c in CostCenter.objects.filter(company=company)}
    job_titles = {j.name.strip(): j for j in JobTitle.objects.filter(company=company)}
    existing_codes = set(Employee.objects.values_list("personnel_code", flat=True))
    existing_ids = set(Employee.objects.values_list("national_id", flat=True))

    errors, created = [], 0
    seen_codes, seen_ids = set(), set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(cell in (None, "") for cell in row):
            continue
        data = {key: (row[i] if i < len(row) else None) for i, (key, _, _) in enumerate(COLUMNS)}

        def fail(message):
            errors.append({"row": row_number, "message": message})

        code = str(data["personnel_code"] or "").strip()
        if code.endswith(".0"):
            code = code[:-2]
        if not code:
            continue  # سطر راهنما یا خالی
        if not str(data["first_name"] or "").strip():
            fail("نام خالی است")
            continue
        if code in existing_codes or code in seen_codes:
            fail(f"کد پرسنلی {code} تکراری است")
            continue

        national_id = str(data["national_id"] or "").strip().replace(".0", "")
        national_id = national_id.zfill(10) if national_id else ""
        if not is_valid_national_id(national_id):
            fail(f"کد ملی «{national_id}» معتبر نیست")
            continue
        if national_id in existing_ids or national_id in seen_ids:
            fail(f"کد ملی {national_id} تکراری است")
            continue

        try:
            hire_date = _parse_jalali(data["hire_date"])
            birth_date = _parse_jalali(data["birth_date"])
        except ValueError as exc:
            fail(str(exc))
            continue
        if hire_date is None:
            fail("تاریخ استخدام خالی است")
            continue

        department = departments.get(str(data["department"] or "").strip())
        cost_center = cost_centers.get(str(data["cost_center"] or "").strip())
        job_title = job_titles.get(str(data["job_title"] or "").strip())
        if create_contracts and not (department and cost_center and job_title):
            missing = []
            if not department:
                missing.append(f"واحد «{data['department']}»")
            if not cost_center:
                missing.append(f"مرکز هزینه «{data['cost_center']}»")
            if not job_title:
                missing.append(f"پست «{data['job_title']}»")
            fail("در سامانه تعریف نشده: " + "، ".join(missing))
            continue

        employee = Employee.objects.create(
            company=company,
            personnel_code=code,
            first_name=str(data["first_name"]).strip(),
            last_name=str(data["last_name"] or "").strip(),
            father_name=str(data["father_name"] or "").strip(),
            national_id=national_id,
            birth_date=birth_date,
            gender=GENDERS.get(str(data["gender"] or "").strip(), "M"),
            marital_status=MARITAL.get(str(data["marital_status"] or "").strip(), "SINGLE"),
            insurance_number=str(data["insurance_number"] or "").strip().replace(".0", ""),
            mobile=str(data["mobile"] or "").strip().replace(".0", ""),
            hire_date=hire_date,
            status=Employee.Status.ACTIVE,
        )
        seen_codes.add(code)
        seen_ids.add(national_id)

        iban = str(data["iban"] or "").strip().upper()
        if iban:
            BankAccount.objects.create(
                employee=employee,
                bank_name=str(data["bank_name"] or "").strip(),
                iban=iban,
                is_default=True,
            )

        if create_contracts:
            EmploymentContract.objects.create(
                employee=employee,
                department=department,
                cost_center=cost_center,
                job_title=job_title,
                effective_from=hire_date,
                base_salary=parse_decimal(data["base_salary"], Decimal("0")),
                seniority_years=int(parse_decimal(data["seniority_years"], Decimal("0"))),
                status=EmploymentContract.Status.ACTIVE,
            )
        created += 1

    if errors:
        # rollback: فایل نیمه‌وارد بدتر از وارد نشدن است.
        # created هم صفر می‌شود، وگرنه گزارش «۱ نفر وارد شد» کنار «هیچ رکوردی
        # ثبت نشد» متناقض و گمراه‌کننده است.
        transaction.set_rollback(True)
        return {"created": 0, "errors": errors, "would_create": created}

    return {"created": created, "errors": []}
