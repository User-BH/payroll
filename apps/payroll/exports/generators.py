"""تولید خروجی‌های قانونی.

هشدار مهم درباره فرمت‌ها:

  * **لیست بیمه (DSK)** طبق ساختار منتشرشدهٔ نرم‌افزار ListDSK ساخته می‌شود ولی
    با یک فایل واقعیِ پذیرفته‌شده تطبیق داده **نشده** است. پیش از اولین ارسال
    رسمی باید با یک فایل واقعی مقایسه شود.
  * **لیست مالیات** اکسل ساختاریافته است، نه فرمت رسمی سامانه مالیاتی.
  * **فایل بانک** عمداً hard-code نشده: ستون‌ها، جداکننده و کدگذاری از تنظیمات
    شرکت خوانده می‌شوند تا تطبیق با بانک بدون تغییر کد ممکن باشد.

جداسازی این سه از هم عمدی است: هر کدام مستقل قابل اصلاح‌اند و اصلاح یکی به
بقیه دست نمی‌زند.
"""

import io
import zipfile
from decimal import Decimal

from django.core.files.base import ContentFile

from apps.payroll.utils import to_jalali

# ---------------------------------------------------------------- کمکی

# cp1256 (عربی ویندوز) حروف مخصوص فارسی «ی» و «ه‌ی چسبان» را ندارد و آن‌ها را
# به «?» تبدیل می‌کند — یعنی نام همه بیمه‌شدگان در فایل DSK خراب می‌شود.
# پیش از کدگذاری، حروف فارسی به معادل عربی‌شان نگاشت می‌شوند؛ این کار استاندارد
# سامانه‌های قدیمی ایرانی است.
LEGACY_MAP = {
    "ی": "ي",  # ی فارسی → ي عربی
    "ک": "ك",  # ک فارسی → ك عربی
    "۰": "0", "۱": "1", "۲": "2", "۳": "3", "۴": "4",
    "۵": "5", "۶": "6", "۷": "7", "۸": "8", "۹": "9",
    "‌": " ",       # نیم‌فاصله → فاصله
    "‏": "", "‎": "",
}
LEGACY_TABLE = str.maketrans(LEGACY_MAP)


def to_legacy(text) -> str:
    """آماده‌سازی متن فارسی برای کدگذاری cp1256."""
    return str(text or "").translate(LEGACY_TABLE)


def encode_legacy(text: str) -> bytes:
    encoded = to_legacy(text).encode("cp1256", errors="replace")
    return encoded


def _payslips(period):
    return (
        period.payslips.select_related("employee", "contract", "contract__job_title")
        .order_by("employee__personnel_code")
    )


def _jalali_compact(value) -> str:
    """تاریخ به شکل ۱۴۰۵۰۵۳۱ که سامانه‌های قدیمی می‌خواهند."""
    jdate = to_jalali(value)
    return f"{jdate.year:04d}{jdate.month:02d}{jdate.day:02d}" if jdate else ""


# ---------------------------------------------------------- لیست بیمه DSK

DSKWOR_FIELDS = [
    "ردیف", "شماره بیمه", "نام", "نام خانوادگی", "نام پدر", "شماره شناسنامه",
    "تاریخ تولد", "محل صدور", "کد ملی", "تاریخ شروع به کار", "تاریخ ترک کار",
    "روزهای کارکرد", "دستمزد روزانه", "دستمزد ماهانه مشمول", "مزایای ماهانه مشمول",
    "جمع دستمزد و مزایای مشمول", "مزایای ماهانه غیرمشمول",
    "حق بیمه سهم بیمه‌شده", "حق بیمه سهم کارفرما", "بیمه بیکاری", "جمع کل حق بیمه",
    "شغل",
]


def generate_insurance(period, company):
    """فایل DSK تأمین اجتماعی — دو فایل متنی DSKKAR و DSKWOR داخل یک zip.

    کدگذاری cp1256 است چون سامانه‌های قدیمی تأمین اجتماعی UTF-8 نمی‌خوانند.

    فقط پرسنل مشمول بیمه در فایل می‌آیند. کسی که «بدون بیمه» علامت خورده و
    قرارداد غیرمشمول، نه در سطرهای فایل هستند، نه در جمع‌ها و نه در تعداد
    نفرات — وگرنه لیست ارسالی با واقعیت نمی‌خواند.
    """
    rows = [p for p in _payslips(period) if p.insurance_applies]

    total_insurable = sum((p.insurable_base for p in rows), Decimal("0"))
    total_employee = sum((p.ins_employee for p in rows), Decimal("0"))
    total_employer = sum((p.ins_employer for p in rows), Decimal("0"))
    total_unemployment = sum(
        (line.amount for p in rows for line in p.lines.all()
         if line.component_code == "UNEMPLOYMENT"),
        Decimal("0"),
    )

    worker_lines = []
    for index, payslip in enumerate(rows, start=1):
        employee = payslip.employee
        daily = (
            payslip.insurable_base / Decimal(payslip.work_days)
            if payslip.work_days else Decimal("0")
        )
        non_insurable = payslip.gross_total - payslip.insurable_base
        worker_lines.append(";".join(str(v) for v in [
            index,
            employee.insurance_number,
            employee.first_name,
            employee.last_name,
            employee.father_name,
            employee.id_number,
            _jalali_compact(employee.birth_date),
            "",
            employee.national_id,
            _jalali_compact(employee.hire_date),
            _jalali_compact(employee.termination_date) if employee.termination_date else "",
            int(payslip.work_days),
            int(daily),
            int(payslip.insurable_base),
            0,
            int(payslip.insurable_base),
            int(non_insurable),
            int(payslip.ins_employee),
            int(payslip.ins_employer),
            int(
                sum((line.amount for line in payslip.lines.all()
                     if line.component_code == "UNEMPLOYMENT"), Decimal("0"))
            ),
            int(payslip.ins_employee + payslip.ins_employer),
            payslip.contract.job_title.name if payslip.contract_id else "",
        ]))

    workshop_line = ";".join(str(v) for v in [
        company.insurance_workshop_code,
        company.name,
        company.address.replace("\n", " ")[:120],
        period.month,
        period.year,
        len(rows),
        int(total_insurable),
        0,
        int(total_insurable),
        int(total_employee),
        int(total_employer),
        int(total_unemployment),
        int(total_employee + total_employer + total_unemployment),
    ])

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("DSKKAR", encode_legacy(workshop_line))
        archive.writestr("DSKWOR", encode_legacy("\r\n".join(worker_lines)))
        # راهنمای ستون‌ها برای بازبینی انسانی؛ سامانه آن را نادیده می‌گیرد
        archive.writestr(
            "COLUMNS.txt",
            ("ترتیب ستون‌های DSKWOR:\n" + "\n".join(
                f"{i}. {name}" for i, name in enumerate(DSKWOR_FIELDS, start=1)
            )).encode("utf-8"),
        )

    return {
        "content": buffer.getvalue(),
        "file_name": f"DSK-{period.year}{period.month:02d}.zip",
        "record_count": len(rows),
        "total_amount": total_employee + total_employer + total_unemployment,
        "summary": {
            "مبنای بیمه": str(int(total_insurable)),
            "سهم بیمه‌شده": str(int(total_employee)),
            "سهم کارفرما": str(int(total_employer)),
            "بیمه بیکاری": str(int(total_unemployment)),
            "کد کارگاه": company.insurance_workshop_code or "ثبت نشده",
        },
    }


# ------------------------------------------------------------ لیست مالیات


def generate_tax(period, company):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = [p for p in _payslips(period) if p.tax_amount > 0]
    total_tax = sum((p.tax_amount for p in rows), Decimal("0"))
    total_base = sum((p.taxable_base for p in rows), Decimal("0"))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "لیست مالیات"
    sheet.sheet_view.rightToLeft = True

    sheet.append([
        f"لیست مالیات حقوق {period.title} — {company.name}"
        f" — پرونده مالیاتی {company.tax_file_number or 'ثبت نشده'}"
    ])
    headers = [
        "ردیف", "نام", "نام خانوادگی", "کد ملی", "شماره پرسنلی", "شغل",
        "روز کارکرد", "ناخالص حقوق و مزایا", "معافیت‌ها و کسورات قانونی",
        "درآمد مشمول مالیات", "مالیات ماه",
    ]
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet.cell(row=1, column=1).font = Font(bold=True, size=12)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    sheet.append(headers)
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F3EEE4")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for index, payslip in enumerate(rows, start=1):
        employee = payslip.employee
        sheet.append([
            index, employee.first_name, employee.last_name, employee.national_id,
            employee.personnel_code,
            payslip.contract.job_title.name if payslip.contract_id else "",
            payslip.work_days,
            payslip.gross_total,
            payslip.gross_total - payslip.taxable_base,
            payslip.taxable_base,
            payslip.tax_amount,
        ])

    sheet.append([])
    sheet.append(["جمع کل", "", "", "", "", "", "", "", "", total_base, total_tax])
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    for column in range(1, len(headers) + 1):
        letter = sheet.cell(row=2, column=column).column_letter
        sheet.column_dimensions[letter].width = max(len(headers[column - 1]) + 3, 13)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return {
        "content": buffer.getvalue(),
        "file_name": f"tax-{period.year}{period.month:02d}.xlsx",
        "record_count": len(rows),
        "total_amount": total_tax,
        "summary": {
            "نفرات مشمول": str(len(rows)),
            "جمع درآمد مشمول": str(int(total_base)),
            "جمع مالیات": str(int(total_tax)),
            "پرونده مالیاتی": company.tax_file_number or "ثبت نشده",
        },
    }


# --------------------------------------------------------- فایل بانک


BANK_TOKENS = {
    "row": lambda i, p: i,
    "personnel_code": lambda i, p: p.employee.personnel_code,
    "national_id": lambda i, p: p.employee.national_id,
    "name": lambda i, p: p.employee.full_name,
    "account": lambda i, p: p.account_snapshot,
    "iban": lambda i, p: p.iban_snapshot,
    "amount": lambda i, p: int(p.net_payable),
    "description": lambda i, p: f"حقوق {p.period.title}",
}


def generate_bank(period, company):
    """فایل پرداخت — ستون‌ها و جداکننده از تنظیمات شرکت."""
    rows = list(_payslips(period))
    total = sum((p.net_payable for p in rows), Decimal("0"))

    columns = [c.strip() for c in (company.bank_file_columns or "").split(",") if c.strip()]
    unknown = [c for c in columns if c not in BANK_TOKENS]
    if unknown:
        raise ValueError(
            "ستون‌های ناشناخته در تنظیمات فایل بانک: " + "، ".join(unknown)
            + ". مقادیر مجاز: " + "، ".join(BANK_TOKENS)
        )
    if not columns:
        raise ValueError("ستون‌های فایل بانک در تنظیمات شرکت تعریف نشده‌اند.")

    delimiter = (company.bank_file_delimiter or ",").replace("\\t", "\t")
    lines = []
    if company.bank_file_include_header:
        lines.append(delimiter.join(columns))
    for index, payslip in enumerate(rows, start=1):
        lines.append(delimiter.join(str(BANK_TOKENS[c](index, payslip)) for c in columns))

    encoding = company.bank_file_encoding or "utf-8-sig"
    text = "\r\n".join(lines) + "\r\n"
    if encoding.lower().replace("-", "") in {"cp1256", "windows1256"}:
        # همان محدودیت DSK: cp1256 «ی» فارسی را ندارد
        text = to_legacy(text)
    content = text.encode(encoding, errors="replace")
    extension = (company.bank_file_extension or "txt").lstrip(".")

    missing = sum(1 for p in rows if not p.account_snapshot)
    return {
        "content": content,
        "file_name": f"bank-{period.year}{period.month:02d}.{extension}",
        "record_count": len(rows),
        "total_amount": total,
        "summary": {
            "بانک": company.name and "رسالت",
            "تعداد رکورد": str(len(rows)),
            "جمع مبلغ": str(int(total)),
            "بدون شماره حساب": str(missing),
            "ستون‌ها": " | ".join(columns),
        },
    }


GENERATORS = {
    "INSURANCE": generate_insurance,
    "TAX": generate_tax,
    "BANK": generate_bank,
}


def generate_export(period, kind, user=None):
    """تولید و ثبت یک خروجی. هر بار تولید، رکورد جدیدی می‌سازد تا تاریخچه بماند."""
    from apps.payroll.models import PayrollExport

    generator = GENERATORS.get(kind)
    if generator is None:
        raise ValueError("نوع خروجی نامعتبر است.")
    if not period.payslips.exists():
        raise ValueError("این دوره فیشی ندارد. ابتدا محاسبه را اجرا کنید.")

    result = generator(period, period.company)
    export = PayrollExport(
        period=period,
        kind=kind,
        file_name=result["file_name"],
        record_count=result["record_count"],
        total_amount=result["total_amount"],
        summary=result["summary"],
        generated_by=user,
    )
    export.file.save(result["file_name"], ContentFile(result["content"]), save=False)
    export.save()
    return export
