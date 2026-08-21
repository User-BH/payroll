import io
from decimal import Decimal

from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from apps.accounts.decorators import can_edit_required, payroll_staff_required
from apps.attendance.models import Timesheet
from apps.employees.models import Employee, EmploymentContract
from apps.org.models import CostCenter, Department
from apps.org.tree import (
    aggregate_period,
    node_key,
    parse_node_key,
    subtree_department_ids,
    walk,
)
from apps.payroll.engine import health
from apps.payroll.engine.health import component_warnings
from apps.payroll.engine.runner import CalculationError, calculate_period
from apps.payroll.exports import generate_export
from apps.payroll.models import (
    CommissionAllocation,
    MonthlyParameter,
    PayrollExport,
    PayrollInput,
    PayrollPeriod,
    Payslip,
)
from apps.payroll.utils import fa_money, jalali_str, parse_decimal

ZERO = Decimal("0")
from apps.payroll_config.models import SalaryComponent


def _period_progress(period):
    """چند نفر از پرسنل فعال کارکرد تأییدشده دارند."""
    if period is None:
        return {"total": 0, "done": 0, "remaining": 0}
    total = Employee.objects.filter(
        company=period.company, status=Employee.Status.ACTIVE
    ).count()
    done = Timesheet.objects.filter(
        period=period, status=Timesheet.Status.APPROVED
    ).count()
    return {"total": total, "done": done, "remaining": max(total - done, 0)}


@payroll_staff_required
def dashboard(request):
    period = PayrollPeriod.objects.select_related("company").order_by("-year", "-month").first()

    progress = _period_progress(period)
    previous = None
    if period:
        previous = (
            PayrollPeriod.objects.filter(company=period.company)
            .exclude(pk=period.pk)
            .filter(status__in=[
                PayrollPeriod.Status.CALCULATED,
                PayrollPeriod.Status.APPROVED,
                PayrollPeriod.Status.LOCKED,
                PayrollPeriod.Status.PAID,
            ])
            .order_by("-year", "-month")
            .first()
        )

    def delta(current, before):
        if not before or not current:
            return None
        change = (current - before) / before * 100
        return round(float(change), 1)

    recent_payslips = []
    if period:
        recent_payslips = list(
            Payslip.objects.filter(period=period)
            .select_related("employee", "department")
            .order_by("-gross_total")[:5]
        )

    # پنل اقلام حقوقی — مثل ماکاپ، بر اساس یک مرکز هزینه فیلتر می‌شود
    cost_center_id = request.GET.get("cc")
    cost_centers = CostCenter.objects.filter(is_active=True).order_by("name")
    selected_cc = None
    if cost_center_id:
        selected_cc = cost_centers.filter(pk=cost_center_id).first()
    if selected_cc is None:
        selected_cc = cost_centers.first()

    components = SalaryComponent.objects.filter(is_active=True).prefetch_related("scopes")
    visible_components = []
    for component in components.order_by("sequence")[:40]:
        scopes = list(component.scopes.all())
        if not scopes:
            visible_components.append(component)
        elif selected_cc and any(
            s.scope_type == "COST_CENTER" and s.scope_id == str(selected_cc.pk) for s in scopes
        ):
            visible_components.append(component)
    visible_components = visible_components[:8]

    return render(
        request,
        "dashboard.html",
        {
            "period": period,
            "progress": progress,
            "recent_payslips": recent_payslips,
            "components": visible_components,
            "cost_centers": cost_centers,
            "selected_cc": selected_cc,
            "active_employees": progress["total"],
            "deltas": {
                "gross": delta(period.total_gross, previous.total_gross) if period and previous else None,
                "net": delta(period.total_net, previous.total_net) if period and previous else None,
                "employer": delta(
                    period.total_employer_cost, previous.total_employer_cost
                ) if period and previous else None,
            },
        },
    )


@payroll_staff_required
def period_list(request):
    periods = (
        PayrollPeriod.objects.select_related("company")
        .annotate(payslip_count=Count("payslips"))
        .order_by("-year", "-month")
    )
    return render(request, "periods/list.html", {"periods": periods})


@can_edit_required
def period_create(request):
    """ساخت دوره حقوقی جدید."""
    from apps.payroll.utils import JALALI_MONTHS, jalali_month_range
    from apps.payroll_config.models import FiscalYear

    company = PayrollPeriod.objects.first().company if PayrollPeriod.objects.exists() else None
    if company is None:
        from apps.org.models import Company

        company = Company.objects.first()

    years = FiscalYear.objects.order_by("-year")
    if request.method == "POST":
        fiscal_year = get_object_or_404(FiscalYear, pk=request.POST.get("fiscal_year"))
        month = int(request.POST.get("month") or 0)
        if not 1 <= month <= 12:
            messages.error(request, "ماه نامعتبر است.")
            return redirect("period_create")
        if PayrollPeriod.objects.filter(
            company=company, year=fiscal_year.year, month=month
        ).exists():
            messages.error(request, "این دوره قبلاً ساخته شده است.")
            return redirect("period_list")

        start, end = jalali_month_range(fiscal_year.year, month)
        period = PayrollPeriod.objects.create(
            company=company,
            fiscal_year=fiscal_year,
            year=fiscal_year.year,
            month=month,
            start_date=start,
            end_date=end,
        )
        messages.success(
            request, f"دوره {period.title} ساخته شد. حالا کارکرد ماهانه را ثبت کنید."
        )
        return redirect("timesheet_grid", pk=period.pk)

    taken = set(PayrollPeriod.objects.values_list("year", "month"))
    return render(
        request,
        "periods/create.html",
        {
            "years": years,
            "months": list(enumerate(JALALI_MONTHS, start=1)),
            "taken": [f"{y}-{m}" for y, m in taken],
        },
    )


@payroll_staff_required
def manual_inputs(request, pk):
    """ثبت مبالغ دستی دوره: مابه التفاوت، حق ماموریت، پورسانت و کسور موردی.

    اقلامی که calc_type آن‌ها MANUAL است مبلغشان از اینجا می‌آید، نه از موتور.
    """
    period = get_object_or_404(PayrollPeriod.objects.select_related("company"), pk=pk)
    components = list(
        SalaryComponent.objects.filter(
            company=period.company, is_active=True, calc_type=SalaryComponent.CalcType.MANUAL
        )
        .prefetch_related("scopes")
        .order_by("sequence")
    )
    selected_code = request.GET.get("component") or (components[0].code if components else None)
    component = next((c for c in components if c.code == selected_code), None)

    query = request.GET.get("q", "").strip()
    contracts = (
        EmploymentContract.objects.filter(
            status=EmploymentContract.Status.ACTIVE,
            employee__company=period.company,
            employee__status=Employee.Status.ACTIVE,
            effective_from__lte=period.end_date,
        )
        .filter(Q(effective_to__isnull=True) | Q(effective_to__gte=period.start_date))
        .select_related("employee", "department", "cost_center", "job_title")
        .order_by("employee__personnel_code")
    )
    if query:
        contracts = contracts.filter(
            Q(employee__first_name__icontains=query)
            | Q(employee__last_name__icontains=query)
            | Q(employee__personnel_code__icontains=query)
        )

    rows = []
    if component:
        amounts = {
            item.employee_id: item.amount
            for item in PayrollInput.objects.filter(period=period, component=component)
        }
        for contract in contracts:
            # قلم فقط به کسانی تعلق می‌گیرد که در دامنه شمولش هستند
            if not component.applies_to(contract):
                continue
            rows.append(
                {
                    "employee": contract.employee,
                    "contract": contract,
                    "amount": amounts.get(contract.employee_id),
                }
            )

    template = (
        "payroll/_manual_rows.html" if request.headers.get("HX-Request") else "payroll/manual.html"
    )
    return render(
        request,
        template,
        {
            "period": period,
            "components": components,
            "component": component,
            "rows": rows,
            "q": query,
            "total": sum((row["amount"] or 0) for row in rows),
        },
    )


@can_edit_required
@require_POST
def manual_input_save(request, pk):
    """ذخیره درجای یک مبلغ دستی — پاسخ HTMX فقط همان سطر است."""
    from apps.payroll.utils import parse_decimal

    period = get_object_or_404(PayrollPeriod, pk=pk)
    if period.is_locked:
        return HttpResponse("دوره قفل است", status=409)

    employee = get_object_or_404(Employee, pk=request.POST.get("employee"))
    component = get_object_or_404(
        SalaryComponent, code=request.POST.get("component"), company=period.company
    )
    amount = parse_decimal(request.POST.get("amount"))

    if amount and amount != 0:
        PayrollInput.objects.update_or_create(
            period=period, employee=employee, component=component,
            defaults={"amount": amount},
        )
    else:
        PayrollInput.objects.filter(
            period=period, employee=employee, component=component
        ).delete()
        amount = None

    return render(
        request,
        "payroll/_manual_row.html",
        {
            "period": period,
            "component": component,
            "row": {"employee": employee, "amount": amount},
            "saved": True,
        },
    )


@payroll_staff_required
def period_detail(request, pk):
    period = get_object_or_404(PayrollPeriod.objects.select_related("company", "fiscal_year"), pk=pk)
    progress = _period_progress(period)
    runs = period.runs.select_related("run_by").order_by("-started_at")[:5]
    # اقلامی که با پیکربندی امروز هیچ عددی تولید نمی‌کنند — پیش از اجرا، نه بعدش.
    # دو دسته‌اند و جدا نشان داده می‌شوند: خرابیِ پیکربندی، و قلمی که سالم است
    # ولی امروز مصداقی ندارد. یکی‌کردنشان یعنی هشدار واقعی زیر اطلاعیه‌ها گم شود.
    all_warnings = component_warnings(period.company, period.legal_parameter)
    warnings = [w for w in all_warnings if w["kind"] == health.BROKEN]
    unreachable = [w for w in all_warnings if w["kind"] == health.UNREACHABLE]

    by_department = (
        Payslip.objects.filter(period=period)
        .values("department__name")
        .annotate(
            count=Count("id"),
            gross=Sum("gross_total"),
            net=Sum("net_payable"),
            tax=Sum("tax_amount"),
            insurance=Sum("ins_employee"),
        )
        .order_by("-gross")
    )

    return render(
        request,
        "periods/detail.html",
        {
            "period": period,
            "progress": progress,
            "runs": runs,
            "by_department": by_department,
            "warnings": warnings,
            "unreachable": unreachable,
            "payslip_count": Payslip.objects.filter(period=period).count(),
        },
    )


@can_edit_required
@require_POST
def period_calculate(request, pk):
    period = get_object_or_404(PayrollPeriod, pk=pk)
    try:
        run = calculate_period(period, user=request.user)
    except CalculationError as exc:
        messages.error(request, str(exc))
    else:
        messages.success(
            request,
            f"محاسبه دوره {period.title} انجام شد — {run.success_count} فیش صادر شد "
            f"(جمع خالص: {fa_money(period.total_net)} ریال).",
        )
    return redirect("period_detail", pk=period.pk)


@can_edit_required
@require_POST
def period_transition(request, pk, action):
    period = get_object_or_404(PayrollPeriod, pk=pk)

    if action == "advance":
        target = period.next_status
        if target is None:
            messages.error(request, "این دوره در آخرین وضعیت است.")
        elif target == PayrollPeriod.Status.CALCULATED and not period.payslips.exists():
            messages.error(request, "ابتدا باید محاسبه دوره اجرا شود.")
        else:
            period.status = target
            if target == PayrollPeriod.Status.APPROVED:
                period.approved_at = timezone.now()
            if target == PayrollPeriod.Status.LOCKED:
                period.locked_at = timezone.now()
                period.locked_by = request.user
            period.save()
            messages.success(request, f"وضعیت دوره به «{period.get_status_display()}» تغییر کرد.")
    elif action == "revert":
        target = period.previous_status
        if target is None:
            messages.error(
                request,
                "برگشت از این وضعیت ممکن نیست. دوره قفل‌شده فقط با فیش اصلاحی تغییر می‌کند.",
            )
        else:
            period.status = target
            period.save()
            messages.info(request, f"وضعیت دوره به «{period.get_status_display()}» برگشت.")
    else:
        messages.error(request, "عملیات نامعتبر است.")

    return redirect("period_detail", pk=period.pk)


@payroll_staff_required
def payslip_list(request, pk):
    period = get_object_or_404(PayrollPeriod, pk=pk)
    query = request.GET.get("q", "").strip()

    payslips = Payslip.objects.filter(period=period).select_related(
        "employee", "department", "cost_center"
    )
    if query:
        payslips = payslips.filter(
            Q(employee__first_name__icontains=query)
            | Q(employee__last_name__icontains=query)
            | Q(employee__personnel_code__icontains=query)
        )
    payslips = payslips.order_by("employee__personnel_code")

    paginator = Paginator(payslips, 30)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    template = "payslips/_rows.html" if request.headers.get("HX-Request") else "payslips/list.html"
    return render(
        request,
        template,
        {"period": period, "page_obj": page_obj, "q": query, "total": payslips.count()},
    )


@payroll_staff_required
def payslip_detail(request, pk):
    payslip = get_object_or_404(
        Payslip.objects.select_related(
            "period", "employee", "contract", "department", "cost_center"
        ),
        pk=pk,
    )
    lines = payslip.lines.select_related("component").order_by("sequence", "id")
    return render(
        request,
        "payslips/detail.html",
        {
            "payslip": payslip,
            "earnings": [l for l in lines if l.kind == SalaryComponent.Kind.EARNING],
            "deductions": [l for l in lines if l.kind == SalaryComponent.Kind.DEDUCTION],
            "employer_costs": [l for l in lines if l.kind == SalaryComponent.Kind.EMPLOYER_COST],
            "bases": [l for l in lines if l.kind == SalaryComponent.Kind.INFO],
        },
    )


def _sheet_data(period):
    """داده‌های برگهٔ محاسبه — مشترک بین صفحهٔ وب و خروجی اکسل.

    سطر: پرسنل · ستون: قلم. ستون‌ها از روی سطرهای واقعی فیش‌ها ساخته می‌شوند نه
    از فهرست اقلام، پس قلمی که این ماه به کسی تعلق نگرفته ستون خالی نمی‌سازد.

    هیچ محاسبهٔ تازه‌ای اینجا انجام نمی‌شود: همهٔ اعداد از `PayslipLine` می‌آیند
    که مبنا، تعداد، ضریب و متن توضیحِ لحظهٔ محاسبه را جدا نگه داشته است. برگه
    یک **نما**ست، نه موتور دوم — وگرنه دو عدد متفاوت می‌داشتیم.
    """
    payslips = list(
        Payslip.objects.filter(period=period)
        .select_related("employee", "department")
        .prefetch_related("lines__component")
        .order_by("employee__personnel_code")
    )

    # ---- ستون‌ها، به ترتیب ترتیبِ محاسبه و گروه‌بندی‌شده بر اساس نوع
    seen = {}
    for payslip in payslips:
        for line in payslip.lines.all():
            seen.setdefault(line.component_code, line)
    ordered = sorted(seen.values(), key=lambda line: (line.sequence, line.component_code))

    Kind = SalaryComponent.Kind
    group_titles = [
        (Kind.EARNING, "استحقاقی", "var(--green)"),
        (Kind.DEDUCTION, "کسورات", "var(--red)"),
        (Kind.EMPLOYER_COST, "هزینه کارفرما", "var(--blue)"),
        (Kind.INFO, "مبناهای محاسبه", "var(--muted)"),
    ]
    groups = []
    columns = []
    for kind, title, color in group_titles:
        members = [line for line in ordered if line.kind == kind]
        if not members:
            continue
        groups.append({"title": title, "color": color, "span": len(members), "kind": kind})
        columns.extend(
            {"code": line.component_code, "name": line.component_name,
             "unit": line.display_unit, "kind": kind}
            for line in members
        )

    # ---- سطرها و جمع ستون‌ها
    rows, totals = [], {column["code"]: ZERO for column in columns}
    for payslip in payslips:
        cells = {}
        for line in payslip.lines.all():
            cells[line.component_code] = line
            totals[line.component_code] = totals.get(line.component_code, ZERO) + line.amount
        # شکل سلول همیشه یکی است، چه آزمایشی در کار باشد چه نه: یک شکل یعنی
        # یک مسیر در قالب و یک مسیر در خروجی اکسل.
        rows.append({
            "payslip": payslip,
            "cells": [{"line": cells.get(c["code"]), "delta": None} for c in columns],
        })

    return {
        "groups": groups,
        "columns": columns,
        "rows": rows,
        "totals": [
            {"value": totals.get(column["code"], ZERO), "delta": None} for column in columns
        ],
        "count": len(payslips),
    }


def _read_simulation(request, period, data):
    """پارامترهای «اجرای آزمایشی» از URL، و تفاوت هر سلول با وضع موجود.

    آزمایش با GET انجام می‌شود نه POST، چون هیچ چیزی تغییر نمی‌دهد و باید
    بشود نتیجه‌اش را با لینک به دیگری نشان داد.
    """
    from apps.payroll.engine.simulate import PARAM_LABELS, TESTABLE_PARAMS, simulate_period

    param_key = request.GET.get("param") or ""
    param_value = (request.GET.get("value") or "").strip()
    component_id = request.GET.get("component") or ""
    component_field = request.GET.get("field") or "rate"
    component_value = (request.GET.get("cvalue") or "").strip()

    params_overrides, component_overrides = {}, {}
    if param_key in PARAM_LABELS and param_value:
        params_overrides[param_key] = param_value
    if component_id and component_value and component_field in ("rate", "fixed_amount"):
        component_overrides[component_id] = {component_field: component_value}

    context = {
        "testable_params": TESTABLE_PARAMS,
        "sim_param": param_key,
        "sim_value": param_value,
        "sim_component": component_id,
        "sim_field": component_field,
        "sim_cvalue": component_value,
        "editable_components": SalaryComponent.objects.filter(
            company=period.company, is_active=True
        ).exclude(kind=SalaryComponent.Kind.INFO).order_by("sequence", "id"),
        "simulation": None,
    }
    if not (params_overrides or component_overrides):
        return context

    result = simulate_period(period, params_overrides, component_overrides)

    # تفاوت هر سلول و هر ستون نسبت به وضع موجود
    for row in data["rows"]:
        employee_id = row["payslip"].employee_id
        for index, column in enumerate(data["columns"]):
            cell = row["cells"][index]
            line = cell["line"]
            current = ZERO
            if line is not None:
                current = line.quantity if line.display_unit in ("DAY", "HOUR") else line.amount
            after = result["cells"].get((employee_id, column["code"]), ZERO)
            cell["delta"] = after - current

    for index, column in enumerate(data["columns"]):
        data["totals"][index]["delta"] = (
            result["totals"].get(column["code"], ZERO) - data["totals"][index]["value"]
        )
    current_net = sum((row["payslip"].net_payable for row in data["rows"]), ZERO)
    new_net = sum((item["net"] for item in result["employees"].values()), ZERO)
    current_cost = sum((row["payslip"].employer_total_cost for row in data["rows"]), ZERO)
    new_cost = sum((item["employer_cost"] for item in result["employees"].values()), ZERO)

    context["simulation"] = {
        "net_before": current_net,
        "net_after": new_net,
        "net_delta": new_net - current_net,
        "cost_before": current_cost,
        "cost_after": new_cost,
        "cost_delta": new_cost - current_cost,
        "param_label": PARAM_LABELS.get(param_key, ""),
        "param_value": param_value,
        "component_changes": result["component_changes"],
    }
    return context


@payroll_staff_required
def period_sheet(request, pk):
    """برگهٔ محاسبهٔ دوره — همان شیت اکسل، ولی هر سلول می‌داند از کجا آمده."""
    period = get_object_or_404(
        PayrollPeriod.objects.select_related("company", "fiscal_year"), pk=pk
    )
    data = _sheet_data(period)
    return render(
        request,
        "periods/sheet.html",
        {"period": period, **data, **_read_simulation(request, period, data)},
    )


@payroll_staff_required
def period_sheet_export(request, pk):
    """همان برگه، عیناً، در اکسل — تا بشود ستون‌به‌ستون با فایل خودشان مقایسه کرد.

    ستون‌ها و ترتیبشان دقیقاً همان صفحهٔ وب است؛ اگر روزی برگه عوض شود این هم
    خودبه‌خود عوض می‌شود، چون هر دو از یک تابع می‌آیند.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    period = get_object_or_404(
        PayrollPeriod.objects.select_related("company"), pk=pk
    )
    data = _sheet_data(period)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "برگه محاسبه"
    sheet.sheet_view.rightToLeft = True

    head = ["کد پرسنلی", "نام و نام خانوادگی", "واحد"] + [c["name"] for c in data["columns"]]
    sheet.append([f"برگه محاسبه {period.title} — {period.company.name}"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(head))
    sheet.cell(row=1, column=1).font = Font(bold=True, size=13)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    sheet.append(head)
    fill = PatternFill("solid", fgColor="F3EEE4")
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row in data["rows"]:
        employee = row["payslip"].employee
        values = [employee.personnel_code, employee.full_name, row["payslip"].department.name]
        for cell in row["cells"]:
            line = cell["line"]
            if line is None:
                values.append(None)
            elif line.display_unit in ("DAY", "HOUR"):
                values.append(line.quantity)
            else:
                values.append(line.amount)
        sheet.append(values)

    total_row = ["", f"جمع {data['count']} نفر", ""] + [t["value"] for t in data["totals"]]
    sheet.append(total_row)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    sheet.freeze_panes = "D3"  # سه ستون هویتی و دو سطر عنوان می‌چسبند
    for index in range(1, len(head) + 1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = 26 if index == 2 else 16

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="sheet-{period.year}-{period.month:02d}.xlsx"'
    )
    workbook.save(response)
    return response


@payroll_staff_required
def payslip_batch_print(request, pk):
    """چاپ دسته‌ای همه فیش‌های یک دوره — هر فیش در یک صفحه A5 جدا.

    فیش فعلی شرکت یک PDF چندصفحه‌ای است؛ چاپ تک‌تک ۱۴۶ فیش عملی نیست.
    """
    period = get_object_or_404(PayrollPeriod.objects.select_related("company"), pk=pk)
    # فیلتر روی «گره چارت» است نه فقط واحد: انتخاب یک مرکز هزینه یعنی فیش همهٔ
    # واحدهای زیرش، وگرنه برای چاپ فیش‌های یک مرکز باید چند بار چاپ گرفت.
    query = request.GET.get("node") or ""
    node = parse_node_key(query)

    payslips = (
        Payslip.objects.filter(period=period)
        .select_related("employee", "contract", "contract__job_title", "department", "cost_center")
        .prefetch_related("lines__component")
        .order_by("employee__personnel_code")
    )
    if node is not None:
        payslips = payslips.filter(department_id__in=subtree_department_ids(node))

    slips = []
    for payslip in payslips:
        lines = [
            line for line in payslip.lines.all()
            if line.component is None or line.component.print_on_payslip
        ]
        slips.append({
            "payslip": payslip,
            "earnings": [l for l in lines if l.kind == SalaryComponent.Kind.EARNING],
            "deductions": [l for l in lines if l.kind == SalaryComponent.Kind.DEDUCTION],
        })

    return render(
        request,
        "payslips/batch.html",
        {
            "period": period,
            "slips": slips,
            "count": len(slips),
            # تورفتگی با فاصلهٔ بدون‌شکست ساخته می‌شود چون <option> نه HTML
            # می‌پذیرد نه فاصلهٔ معمولی را نگه می‌دارد.
            "nodes": [
                {"key": node_key(item), "indent": "\u00a0\u00a0\u00a0" * depth,
                 "name": item.name, "is_cost_center": isinstance(item, CostCenter)}
                for item, depth in walk(period.company)
            ],
            "selected_node": query,
            "node": node,
        },
    )


@payroll_staff_required
def payslip_print(request, pk):
    payslip = get_object_or_404(
        Payslip.objects.select_related("period", "employee", "contract", "department"), pk=pk
    )
    # تیک «چاپ در فیش» فقط روی نسخهٔ چاپی اثر دارد؛ صفحهٔ داخلی همه را نشان می‌دهد.
    lines = [
        line
        for line in payslip.lines.select_related("component").order_by("sequence", "id")
        if line.component is None or line.component.print_on_payslip
    ]
    return render(
        request,
        "payslips/print.html",
        {
            "payslip": payslip,
            "earnings": [l for l in lines if l.kind == SalaryComponent.Kind.EARNING],
            "deductions": [l for l in lines if l.kind == SalaryComponent.Kind.DEDUCTION],
        },
    )


@payroll_staff_required
def dispute_list(request):
    """اعتراض‌های ثبت‌شده پرسنل روی فیش‌ها."""
    open_disputes = (
        Payslip.objects.filter(ack_status=Payslip.Ack.DISPUTED)
        .select_related("employee", "period", "department")
        .order_by("-disputed_at")
    )
    resolved = (
        Payslip.objects.filter(ack_status=Payslip.Ack.RESOLVED)
        .select_related("employee", "period", "reviewed_by")
        .order_by("-reviewed_at")[:20]
    )
    return render(
        request,
        "disputes/list.html",
        {"open_disputes": open_disputes, "resolved": resolved},
    )


@can_edit_required
@require_POST
def dispute_resolve(request, pk):
    payslip = get_object_or_404(Payslip, pk=pk)
    note = (request.POST.get("note") or "").strip()
    if len(note) < 5:
        messages.error(request, "پاسخ به اعتراض نمی‌تواند خالی باشد.")
        return redirect("dispute_list")

    payslip.ack_status = Payslip.Ack.RESOLVED
    payslip.review_note = note
    payslip.reviewed_by = request.user
    payslip.reviewed_at = timezone.now()
    payslip.save_ack(["ack_status", "review_note", "reviewed_by", "reviewed_at"])
    messages.success(
        request,
        f"پاسخ برای {payslip.employee.full_name} ثبت شد و در پرتال او نمایش داده می‌شود.",
    )
    return redirect("dispute_list")


# ----------------------------------------------------------------- گزارش‌ها


@payroll_staff_required
def reports(request):
    periods = (
        PayrollPeriod.objects.annotate(payslip_count=Count("payslips"))
        .filter(payslip_count__gt=0)
        .order_by("-year", "-month")
    )
    return render(request, "reports/index.html", {"periods": periods})


def _report_rows(period, node=None):
    """سطرهای گزارش، در صورت انتخاب گره فقط زیردرخت همان گره.

    فیلتر روی زیردرخت است نه روی خود گره: انتخاب یک مرکز هزینه باید همهٔ
    واحدهای زیرش را بیاورد، وگرنه «لیست بیمهٔ فروش» ناقص چاپ می‌شود.
    """
    rows = (
        Payslip.objects.filter(period=period)
        .select_related("employee", "department")
        .order_by("employee__personnel_code")
    )
    if node is not None:
        rows = rows.filter(department_id__in=subtree_department_ids(node))
    return rows


def _node_filter(request, period):
    """گره انتخاب‌شده و فهرست گره‌ها برای نوار فیلتر گزارش‌ها."""
    key = request.GET.get("node") or ""
    node = parse_node_key(key)
    return node, {
        "node": node,
        "node_key": key,
        "nodes": [
            {"key": node_key(item), "indent": "\u00a0\u00a0\u00a0" * depth,
             "name": item.name, "is_cost_center": isinstance(item, CostCenter)}
            for item, depth in walk(period.company)
        ],
    }


@payroll_staff_required
def report_insurance(request, pk):
    """لیست بیمه — فقط پرسنل مشمول.

    پرسنلِ بدون بیمه و قرارداد غیرمشمول اینجا نمی‌آیند و در تعداد نفرات هم شمرده
    نمی‌شوند، دقیقاً مثل فایل خروجی تأمین اجتماعی. تعدادشان جداگانه گزارش
    می‌شود تا «چرا ۱۴۵ نفر شد نه ۱۴۶؟» بی‌جواب نماند.
    """
    period = get_object_or_404(PayrollPeriod.objects.select_related("company"), pk=pk)
    node, node_context = _node_filter(request, period)
    all_rows = _report_rows(period, node)
    rows = all_rows.filter(insurance_applies=True)
    excluded = all_rows.filter(insurance_applies=False)
    totals = rows.aggregate(
        insurable=Sum("insurable_base"),
        employee=Sum("ins_employee"),
        employer=Sum("ins_employer"),
    )
    return render(
        request,
        "reports/insurance.html",
        {
            "period": period,
            "rows": rows,
            "totals": totals,
            "count": rows.count(),
            "excluded": excluded,
            "excluded_count": excluded.count(),
            **node_context,
        },
    )


@payroll_staff_required
def report_tax(request, pk):
    period = get_object_or_404(PayrollPeriod.objects.select_related("company"), pk=pk)
    node, node_context = _node_filter(request, period)
    rows = _report_rows(period, node).filter(tax_amount__gt=0)
    totals = rows.aggregate(taxable=Sum("taxable_base"), tax=Sum("tax_amount"))
    return render(
        request,
        "reports/tax.html",
        {"period": period, "rows": rows, "totals": totals, "count": rows.count(),
         **node_context},
    )


@payroll_staff_required
def report_bank(request, pk):
    period = get_object_or_404(PayrollPeriod.objects.select_related("company"), pk=pk)
    node, node_context = _node_filter(request, period)
    rows = _report_rows(period, node)
    totals = rows.aggregate(net=Sum("net_payable"))
    missing_account = rows.filter(account_snapshot="").count()
    return render(
        request,
        "reports/bank.html",
        {
            "period": period,
            "rows": rows,
            "totals": totals,
            "count": rows.count(),
            "missing_account": missing_account,
            **node_context,
        },
    )


@payroll_staff_required
def report_org(request, pk):
    """گزارش تجمیعی روی چارت سازمانی — بند ۸ سند.

    هر سطر جمعِ زیردرخت خودش است، پس جمع سطرهای سطح اول باید دقیقاً برابر جمع
    کل دوره شود. اگر فیشی به واحدی وصل باشد که دیگر در چارت نیست، در سطر جدا
    می‌آید تا این تساوی هیچ‌وقت بی‌سروصدا بشکند.
    """
    period = get_object_or_404(PayrollPeriod.objects.select_related("company"), pk=pk)
    rows, orphan, grand = aggregate_period(period)
    return render(
        request,
        "reports/org.html",
        {
            "period": period,
            "rows": rows,
            "orphan": orphan,
            "grand": grand,
            "unassigned": Department.objects.filter(
                company=period.company, cost_center__isnull=True, parent__isnull=True
            ),
        },
    )


@payroll_staff_required
def export_center(request, pk):
    """مرکز خروجی‌های قانونی یک دوره."""
    period = get_object_or_404(PayrollPeriod.objects.select_related("company"), pk=pk)
    company = period.company
    exports = period.exports.select_related("generated_by").order_by("-generated_at")

    latest = {}
    for export in exports:
        latest.setdefault(export.kind, export)

    warnings = []
    if not company.insurance_workshop_code:
        warnings.append("کد کارگاه بیمه در اطلاعات شرکت ثبت نشده — لیست بیمه ناقص می‌شود.")
    if not company.tax_file_number:
        warnings.append("شماره پرونده مالیاتی ثبت نشده — لیست مالیات ناقص می‌شود.")
    missing_account = period.payslips.filter(account_snapshot="").count()
    if missing_account:
        warnings.append(
            f"{missing_account} نفر شماره حساب ندارند و در فایل بانک ناقص می‌آیند."
        )

    return render(
        request,
        "exports/center.html",
        {
            "period": period,
            "company": company,
            "exports": exports,
            "latest": latest,
            "kinds": PayrollExport.Kind.choices,
            "warnings": warnings,
            "payslip_count": period.payslips.count(),
        },
    )


@can_edit_required
@require_POST
def export_generate(request, pk, kind):
    period = get_object_or_404(PayrollPeriod, pk=pk)
    try:
        export = generate_export(period, kind, user=request.user)
    except ValueError as exc:
        messages.error(request, str(exc))
    else:
        messages.success(
            request,
            f"{export.get_kind_display()} تولید شد — {export.record_count} رکورد. "
            "پیش از ارسال رسمی، محتوایش را بازبینی کنید.",
        )
    return redirect("export_center", pk=period.pk)


@payroll_staff_required
def export_download(request, pk):
    export = get_object_or_404(PayrollExport.objects.select_related("period"), pk=pk)
    response = HttpResponse(export.file.read(), content_type="application/octet-stream")
    response["Content-Disposition"] = f'attachment; filename="{export.file_name}"'
    return response


@can_edit_required
@require_POST
def export_update_status(request, pk):
    """ثبت کد رهگیری و وضعیت ارسال."""
    export = get_object_or_404(PayrollExport.objects.select_related("period"), pk=pk)
    status = request.POST.get("status")
    if status not in dict(PayrollExport.Status.choices):
        messages.error(request, "وضعیت نامعتبر است.")
        return redirect("export_center", pk=export.period.pk)

    tracking = (request.POST.get("tracking_code") or "").strip()
    if status in {PayrollExport.Status.SUBMITTED, PayrollExport.Status.ACCEPTED} and not tracking:
        messages.error(request, "برای ثبت ارسال، کد رهگیری اجباری است.")
        return redirect("export_center", pk=export.period.pk)

    export.status = status
    export.tracking_code = tracking
    export.note = (request.POST.get("note") or "").strip()[:250]
    if status != PayrollExport.Status.GENERATED and not export.submitted_at:
        export.submitted_at = timezone.now()
    export.save()
    messages.success(request, "وضعیت خروجی ثبت شد.")
    return redirect("export_center", pk=export.period.pk)


REPORT_SPECS = {
    "insurance": {
        "title": "لیست بیمه",
        "headers": [
            "ردیف", "کد پرسنلی", "نام", "نام خانوادگی", "کد ملی", "شماره بیمه",
            "روز کارکرد", "مبنای بیمه", "سهم کارگر", "سهم کارفرما",
        ],
        "fields": lambda p, i: [
            i,
            p.employee.personnel_code,
            p.employee.first_name,
            p.employee.last_name,
            p.employee.national_id,
            p.employee.insurance_number,
            p.work_days,
            p.insurable_base,
            p.ins_employee,
            p.ins_employer,
        ],
    },
    "tax": {
        "title": "لیست مالیات",
        "headers": [
            "ردیف", "کد پرسنلی", "نام و نام خانوادگی", "کد ملی",
            "ناخالص", "مبنای مالیات", "مالیات",
        ],
        "fields": lambda p, i: [
            i,
            p.employee.personnel_code,
            p.employee.full_name,
            p.employee.national_id,
            p.gross_total,
            p.taxable_base,
            p.tax_amount,
        ],
    },
    "bank": {
        "title": "فایل پرداخت بانک",
        # پرداخت این شرکت از طریق بانک رسالت و بر مبنای شماره حساب است، نه شبا
        "headers": ["ردیف", "کد پرسنلی", "نام و نام خانوادگی", "بانک", "شماره حساب", "مبلغ خالص"],
        "fields": lambda p, i: [
            i,
            p.employee.personnel_code,
            p.employee.full_name,
            p.bank_snapshot or "رسالت",
            p.account_snapshot,
            p.net_payable,
        ],
    },
    "payslips": {
        "title": "خلاصه فیش‌ها",
        "headers": [
            "ردیف", "کد پرسنلی", "نام و نام خانوادگی", "واحد", "روز کارکرد",
            "ناخالص", "مبنای بیمه", "مبنای مالیات", "بیمه", "مالیات", "خالص",
        ],
        "fields": lambda p, i: [
            i,
            p.employee.personnel_code,
            p.employee.full_name,
            p.department.name,
            p.work_days,
            p.gross_total,
            p.insurable_base,
            p.taxable_base,
            p.ins_employee,
            p.tax_amount,
            p.net_payable,
        ],
    },
}


@payroll_staff_required
def report_export(request, pk, kind):
    """خروجی اکسل برای مرور و بایگانی داخلی.

    خروجی‌های رسمی با ثبت کد رهگیری، در «مرکز خروجی‌های قانونی» تولید می‌شوند.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    spec = REPORT_SPECS.get(kind)
    if spec is None:
        messages.error(request, "نوع گزارش نامعتبر است.")
        return redirect("reports")

    period = get_object_or_404(PayrollPeriod.objects.select_related("company"), pk=pk)
    # همان گرهٔ صفحهٔ گزارش، تا فایل اکسل با چیزی که کاربر روی صفحه می‌بیند یکی
    # باشد؛ خروجیِ کاملِ بی‌خبر، بدترین حالت است.
    node = parse_node_key(request.GET.get("node"))
    rows = _report_rows(period, node)
    # همان صافیِ صفحهٔ گزارش روی فایل هم اعمال می‌شود: «لیست بیمه»‌ای که
    # پرسنلِ بدون بیمه و قرارداد غیرمشمول را هم داشته باشد با صفحه و با فایل تأمین
    # اجتماعی نمی‌خواند.
    if kind == "insurance":
        rows = rows.filter(insurance_applies=True)
    elif kind == "tax":
        rows = rows.filter(tax_amount__gt=0)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec["title"][:30]
    sheet.sheet_view.rightToLeft = True

    scope = f" — {node.full_path}" if node is not None else ""
    sheet.append([f"{spec['title']} — {period.title} — {period.company.name}{scope}"])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(spec["headers"]))
    sheet.cell(row=1, column=1).font = Font(bold=True, size=13)
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    sheet.append(spec["headers"])
    header_fill = PatternFill("solid", fgColor="F3EEE4")
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for index, payslip in enumerate(rows, start=1):
        sheet.append(spec["fields"](payslip, index))

    # عرض ستون‌ها بر اساس محتوا. سطر اول merge شده است و سلول‌های merge شده
    # column_letter ندارند، پس از شماره ستون حرف را می‌سازیم.
    for index in range(1, len(spec["headers"]) + 1):
        letter = get_column_letter(index)
        width = max(
            (len(str(sheet.cell(row=row, column=index).value or "")) for row in range(2, sheet.max_row + 1)),
            default=10,
        )
        sheet.column_dimensions[letter].width = min(max(width + 3, 12), 42)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"{kind}-{period.year}-{period.month:02d}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ------------------------------------------------ آیتم‌های محاسباتی ماهانه


@payroll_staff_required
def monthly_parameters(request, pk):
    """مقادیر محاسباتی‌ای که هر ماه فرق می‌کنند.

    آنچه سالانه ثابت است در «تنظیمات سال مالی» می‌ماند. اینجا فقط استثنای همان
    ماه ثبت می‌شود: خالی گذاشتن هر خانه یعنی «همان مقدار سالانه».
    """
    from apps.payroll.engine.runner import resolve_legal_parameter
    from apps.payroll_config.monthly import annual_value, grouped_params

    period = get_object_or_404(
        PayrollPeriod.objects.select_related("company", "fiscal_year"), pk=pk
    )
    try:
        legal_parameter = resolve_legal_parameter(period)
    except Exception:  # noqa: BLE001 — نبود پارامتر سال باید پیام بدهد نه ۵۰۰
        messages.error(
            request,
            f"برای سال مالی {period.fiscal_year.year} پارامتر قانونی تعریف نشده است. "
            "ابتدا «تنظیمات سال مالی» را کامل کنید.",
        )
        return redirect("period_detail", pk=period.pk)

    saved = {
        item.key: item for item in MonthlyParameter.objects.filter(period=period)
    }

    groups = []
    for group_name, rows in grouped_params().items():
        items = []
        for row in rows:
            item = saved.get(row["key"])
            items.append({
                **row,
                "annual": annual_value(legal_parameter, row["key"]),
                "monthly": item.value if item else None,
                # توضیحِ ذخیره‌شدهٔ کاربر، نه راهنمای رجیستری — راهنما زیر
                # عنوان نشان داده می‌شود و نباید داخل خانهٔ ورودی بنشیند.
                "note": item.note if item else "",
                "hint": row["note"],
            })
        groups.append({"name": group_name, "items": items})

    return render(
        request,
        "settings/monthly_parameters.html",
        {
            "period": period,
            "groups": groups,
            "legal_parameter": legal_parameter,
            "override_count": len(saved),
        },
    )


@can_edit_required
@require_POST
def monthly_parameters_save(request, pk):
    """ذخیره مقادیر ماهانه. خانهٔ خالی یعنی حذف استثنا و برگشت به مقدار سالانه."""
    from apps.payroll_config.monthly import PARAM_KEYS

    period = get_object_or_404(PayrollPeriod, pk=pk)
    if period.is_locked:
        messages.error(request, "دوره قفل شده است و پارامترهایش قابل تغییر نیستند.")
        return redirect("monthly_parameters", pk=period.pk)

    created, removed = 0, 0
    for key in PARAM_KEYS:
        raw = (request.POST.get(f"value__{key}") or "").strip()
        note = (request.POST.get(f"note__{key}") or "").strip()[:160]
        if not raw:
            removed += MonthlyParameter.objects.filter(period=period, key=key).delete()[0]
            continue
        value = parse_decimal(raw, None)
        if value is None or value < 0:
            messages.error(request, f"مقدار «{raw}» برای این ماه معتبر نیست.")
            return redirect("monthly_parameters", pk=period.pk)
        MonthlyParameter.objects.update_or_create(
            period=period, key=key,
            defaults={"value": value, "note": note, "updated_by": request.user},
        )
        created += 1

    messages.success(
        request,
        f"{created} مقدار ماهانه برای {period.title} ثبت شد"
        + (f" و {removed} مورد به مقدار سالانه برگشت." if removed else ".")
        + " برای اعمال روی فیش‌ها، دوره را دوباره محاسبه کنید.",
    )
    return redirect("monthly_parameters", pk=period.pk)


# ------------------------------------------------------- تخصیص پورسانت


@payroll_staff_required
def commission_allocations(request, pk):
    """تخصیص پورسانت به حق مأموریت و اضافه‌کاری.

    هر سطر نشان می‌دهد پورسانت اولیه چقدر بوده، چقدر کجا رفته و چقدر مانده —
    و اینکه ظرفیت هر کدام در همین ماه چقدر است.
    """
    from apps.payroll.commission import build_plan, commission_totals
    from apps.payroll.engine.runner import (
        resolve_effective_params, resolve_legal_parameter,
    )

    period = get_object_or_404(
        PayrollPeriod.objects.select_related("company", "fiscal_year"), pk=pk
    )
    try:
        params = resolve_effective_params(period, resolve_legal_parameter(period))
    except Exception:  # noqa: BLE001
        messages.error(request, "پارامتر قانونی این سال تعریف نشده است.")
        return redirect("period_detail", pk=period.pk)

    totals = commission_totals(period)
    timesheets = {ts.employee_id: ts for ts in Timesheet.objects.filter(period=period)}
    saved = {a.employee_id: a for a in CommissionAllocation.objects.filter(period=period)}

    employees = {
        e.pk: e for e in Employee.objects.filter(pk__in=totals.keys())
    }

    rows = []
    for employee_id, source in sorted(
        totals.items(), key=lambda kv: employees[kv[0]].personnel_code if kv[0] in employees else ""
    ):
        employee = employees.get(employee_id)
        if employee is None:
            continue
        allocation = saved.get(employee_id)
        plan = build_plan(period, employee, params, source, timesheets.get(employee_id))
        if allocation is not None:
            plan.manual(allocation.to_mission, allocation.to_overtime)
        else:
            plan.auto()
        rows.append({
            "employee": employee,
            "source": source,
            "plan": plan,
            "allocation": allocation,
            "is_manual": bool(allocation and allocation.is_manual),
        })

    return render(
        request,
        "payroll/commission.html",
        {
            "period": period,
            "rows": rows,
            "mission_rate": params.mission_daily_rate,
            "mission_max_days": params.mission_max_days,
            "overtime_rate": params.overtime_hourly_rate,
            "overtime_max_hours": params.overtime_max_hours,
            "totals": {
                "source": sum((r["source"] for r in rows), 0),
                "mission": sum((r["plan"].to_mission for r in rows), 0),
                "overtime": sum((r["plan"].to_overtime for r in rows), 0),
                "remaining": sum((r["plan"].remaining for r in rows), 0),
            },
        },
    )


@can_edit_required
@require_POST
def commission_allocations_save(request, pk):
    """ذخیره تقسیم دستی، یا برگرداندن همه به تقسیم خودکار."""
    from apps.payroll.commission import build_plan, commission_totals
    from apps.payroll.engine.runner import (
        resolve_effective_params, resolve_legal_parameter,
    )

    period = get_object_or_404(PayrollPeriod, pk=pk)
    if period.is_locked:
        messages.error(request, "دوره قفل شده است.")
        return redirect("commission_allocations", pk=period.pk)

    params = resolve_effective_params(period, resolve_legal_parameter(period))
    totals = commission_totals(period)
    timesheets = {ts.employee_id: ts for ts in Timesheet.objects.filter(period=period)}

    if request.POST.get("action") == "auto":
        CommissionAllocation.objects.filter(period=period).delete()
        messages.success(
            request,
            "همهٔ تخصیص‌ها به تقسیم خودکار برگشتند. "
            "برای اعمال روی فیش‌ها، دوره را دوباره محاسبه کنید.",
        )
        return redirect("commission_allocations", pk=period.pk)

    changed = 0
    for employee in Employee.objects.filter(pk__in=totals.keys()):
        prefix = f"emp__{employee.pk}"
        if f"{prefix}__mission" not in request.POST:
            continue
        plan = build_plan(
            period, employee, params, totals[employee.pk], timesheets.get(employee.pk)
        ).manual(
            parse_decimal(request.POST.get(f"{prefix}__mission"), 0),
            parse_decimal(request.POST.get(f"{prefix}__overtime"), 0),
        )
        CommissionAllocation.objects.update_or_create(
            period=period, employee=employee,
            defaults={
                **plan.as_fields(),
                "is_manual": True,
                "note": (request.POST.get(f"{prefix}__note") or "").strip()[:160],
                "updated_by": request.user,
            },
        )
        changed += 1

    messages.success(
        request,
        f"تخصیص پورسانت {changed} نفر ثبت شد. مبالغ خارج از ظرفیت به سقف مجاز "
        "محدود شدند و مازادشان در ماندهٔ پورسانت ماند. "
        "برای اعمال روی فیش‌ها، دوره را دوباره محاسبه کنید.",
    )
    return redirect("commission_allocations", pk=period.pk)
