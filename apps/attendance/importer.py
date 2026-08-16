"""ورود کارکرد ماهانه از اکسل.

ثبت دستی کارکرد ۱۴۶ نفر، هر ماه، شدنی نیست. مثل ورود پرسنل، اینجا هم قاعده «یا
همه یا هیچ» برقرار است: اگر یک سطر ایراد داشته باشد، هیچ ردیفی به‌روز نمی‌شود.

تفاوت مهم با ورود پرسنل: اینجا رکورد جدید ساخته نمی‌شود. کارکرد فقط برای
پرسنلی که در همان دوره قرارداد فعال دارند به‌روزرسانی می‌شود؛ کد پرسنلی ناشناس
خطاست، نه ساخت رکورد جدید.
"""

import io
from decimal import Decimal

from django.db import transaction

from apps.attendance.leave import DEFAULT_DAY_MINUTES
from apps.attendance.models import Timesheet, TimesheetImport
from apps.attendance.services import minutes_from_parts
from apps.payroll.utils import parse_decimal

COLUMNS = [
    ("personnel_code", "کد پرسنلی", True),
    ("full_name", "نام و نام خانوادگی (فقط برای کنترل)", False),
    ("work_days", "روز کارکرد", True),
    ("leave_m", "مرخصی — دقیقه", False),
    ("leave_h", "مرخصی — ساعت", False),
    ("leave_d", "مرخصی — روز", False),
    ("absence_days", "غیبت (روز)", False),
    ("unpaid_leave_days", "مرخصی بدون حقوق (روز)", False),
    ("sick_leave_days", "استعلاجی (روز)", False),
    ("mission_days", "حق مأموریت (روز)", False),
    ("ot_m", "اضافه‌کاری — دقیقه", False),
    ("ot_h", "اضافه‌کاری — ساعت", False),
    ("ot_d", "اضافه‌کاری — روز", False),
    ("holiday_hours", "تعطیل‌کاری (ساعت)", False),
]

# شب‌کاری و جمعه‌کاری از فایل کارکرد برداشته شده‌اند؛ اضافه‌کاری مثل مرخصی
# سه‌ستونی (دقیقه/ساعت/روز) است و به دقیقه ذخیره می‌شود.
DECIMAL_FIELDS = [
    "work_days", "absence_days", "unpaid_leave_days", "sick_leave_days",
    "mission_days", "holiday_hours",
]


def build_template(period) -> bytes:
    """فایل الگو، از پیش پر شده با پرسنل همان دوره.

    کاربر فقط اعداد را عوض می‌کند؛ کد پرسنلی از قبل درست است و احتمال خطای
    تایپی از بین می‌رود.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"کارکرد {period.month}-{period.year}"
    sheet.sheet_view.rightToLeft = True

    sheet.append([label for _, label, _ in COLUMNS])
    fill = PatternFill("solid", fgColor="F3EEE4")
    for index, (_, _, required) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True, color="A81812" if required else "1B1815")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        sheet.column_dimensions[cell.column_letter].width = max(len(cell.value) + 3, 13)
    sheet.row_dimensions[1].height = 40
    sheet.freeze_panes = "A2"

    rows = (
        Timesheet.objects.filter(period=period)
        .select_related("employee")
        .order_by("employee__personnel_code")
    )
    for timesheet in rows:
        leave = timesheet.leave_parts
        overtime = timesheet.overtime_parts
        sheet.append([
            timesheet.employee.personnel_code,
            timesheet.employee.full_name,
            timesheet.work_days,
            leave["m"], leave["h"], leave["d"],
            timesheet.absence_days,
            timesheet.unpaid_leave_days,
            timesheet.sick_leave_days,
            timesheet.mission_days,
            overtime["m"], overtime["h"], overtime["d"],
            timesheet.holiday_hours,
        ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


@transaction.atomic
def import_timesheets(file_obj, period, user=None, file_name="") -> dict:
    """خواندن فایل و به‌روزرسانی کارکرد دوره. یا همه یا هیچ."""
    from openpyxl import load_workbook

    if not period.is_editable:
        return {
            "updated": 0,
            "errors": [{"row": 0, "message": "دوره در وضعیت پیش‌نویس نیست و کارکردش قفل است."}],
        }

    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook.active

    day_minutes = (
        period.legal_parameter.leave_day_minutes
        if period.legal_parameter
        else DEFAULT_DAY_MINUTES
    )
    sheets = {
        ts.employee.personnel_code: ts
        for ts in Timesheet.objects.filter(period=period).select_related("employee")
    }

    errors, updated, seen = [], 0, set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(cell in (None, "") for cell in row):
            continue
        data = {key: (row[i] if i < len(row) else None) for i, (key, _, _) in enumerate(COLUMNS)}

        code = str(data["personnel_code"] or "").strip()
        if code.endswith(".0"):
            code = code[:-2]
        if not code:
            continue

        timesheet = sheets.get(code)
        if timesheet is None:
            errors.append({
                "row": row_number,
                "message": f"کد پرسنلی {code} در این دوره کارکرد ندارد "
                           "(یا قرارداد فعال ندارد، یا کد اشتباه است)",
            })
            continue
        if code in seen:
            errors.append({"row": row_number, "message": f"کد پرسنلی {code} تکراری است"})
            continue
        seen.add(code)

        work_days = parse_decimal(data["work_days"], None)
        if work_days is None or work_days < 0 or work_days > 31:
            errors.append({
                "row": row_number,
                "message": f"روز کارکرد «{data['work_days']}» معتبر نیست (باید بین ۰ و ۳۱ باشد)",
            })
            continue

        for field in DECIMAL_FIELDS:
            value = parse_decimal(data[field])
            if value < 0:
                errors.append({"row": row_number, "message": f"«{field}» نمی‌تواند منفی باشد"})
                break
            setattr(timesheet, field, value)
        else:
            timesheet.leave_minutes = minutes_from_parts(data, "leave", day_minutes)
            timesheet.overtime_minutes = minutes_from_parts(data, "ot", day_minutes)
            timesheet.source = Timesheet.Source.IMPORT
            timesheet.entered_by = user
            timesheet.save()
            updated += 1

    log = TimesheetImport.objects.create(
        period=period,
        file_name=file_name or "بدون نام",
        row_count=updated + len(errors),
        error_count=len(errors),
        log=errors[:200],
        imported_by=user,
    )

    if errors:
        # rollback کامل — شامل خودِ رکورد لاگ، چون واردکردنی انجام نشده است
        transaction.set_rollback(True)
        return {"updated": 0, "errors": errors, "would_update": updated}

    return {"updated": updated, "errors": [], "log_id": log.pk}
