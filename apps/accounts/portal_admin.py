"""ساخت حساب پرتال برای پرسنل.

رمز اولیه به‌صورت تصادفی ساخته و **فقط یک بار** نمایش داده می‌شود تا چاپ و
تحویل شود. چون این رمز روی کاغذ دست به دست می‌گردد، حساب با پرچم
must_change_password ساخته می‌شود و پرسنل در اولین ورود مجبور به تغییر آن است.
"""

import io
import secrets
import string

from django.contrib import messages
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from apps.accounts.decorators import can_edit_required, payroll_staff_required
from apps.employees.models import Employee

User = get_user_model()

# حروف مبهم (l، I، O، 0، 1) حذف شده‌اند تا رمز روی کاغذ اشتباه خوانده نشود
ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnpqrstuvwxyz23456789"


def make_password_value(length=10) -> str:
    return "".join(secrets.choice(ALPHABET) for _ in range(length))


@payroll_staff_required
def portal_accounts(request):
    employees = (
        Employee.objects.filter(status=Employee.Status.ACTIVE)
        .select_related("user")
        .order_by("personnel_code")
    )
    without = [e for e in employees if e.user_id is None]
    return render(
        request,
        "accounts/portal_accounts.html",
        {
            "employees": employees,
            "without_count": len(without),
            "with_count": len(employees) - len(without),
            "created": request.session.pop("portal_created", None),
        },
    )


@can_edit_required
@require_POST
@transaction.atomic
def portal_accounts_create(request):
    """ساخت حساب برای همه پرسنل فعالی که هنوز حساب ندارند."""
    targets = (
        Employee.objects.filter(status=Employee.Status.ACTIVE, user__isnull=True)
        .order_by("personnel_code")
    )
    if not targets.exists():
        messages.info(request, "همه پرسنل فعال از قبل حساب پرتال دارند.")
        return redirect("portal_accounts")

    created = []
    for employee in targets:
        username = employee.personnel_code
        if User.objects.filter(username=username).exists():
            # نام کاربری اشغال است؛ رد می‌شود تا حساب اشتباهی به کسی وصل نشود
            continue
        password = make_password_value()
        user = User.objects.create(
            username=username,
            first_name=employee.first_name,
            last_name=employee.last_name,
            role=User.Role.EMPLOYEE,
            mobile=employee.mobile,
            must_change_password=True,
        )
        user.set_password(password)
        user.save()
        employee.user = user
        employee.save(update_fields=["user"])
        created.append({
            "code": employee.personnel_code,
            "name": employee.full_name,
            "password": password,
        })

    request.session["portal_created"] = created
    messages.success(
        request,
        f"{len(created)} حساب ساخته شد. رمزها فقط همین یک بار نمایش داده می‌شوند — "
        "فهرست را دانلود یا چاپ کنید.",
    )
    return redirect("portal_accounts")


@can_edit_required
@require_POST
def portal_account_reset(request, pk):
    """بازنشانی رمز یک نفر."""
    employee = get_object_or_404(Employee.objects.select_related("user"), pk=pk)
    if employee.user is None:
        messages.error(request, "این پرسنل حساب پرتال ندارد.")
        return redirect("portal_accounts")

    password = make_password_value()
    employee.user.set_password(password)
    employee.user.must_change_password = True
    employee.user.save()
    request.session["portal_created"] = [{
        "code": employee.personnel_code,
        "name": employee.full_name,
        "password": password,
    }]
    messages.success(request, f"رمز {employee.full_name} بازنشانی شد.")
    return redirect("portal_accounts")


@payroll_staff_required
def portal_accounts_export(request):
    """خروجی اکسل رمزهای تازه‌ساخته‌شده برای چاپ و تحویل."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    created = request.session.get("portal_created") or []
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "حساب پرتال"
    sheet.sheet_view.rightToLeft = True

    sheet.append(["کد پرسنلی", "نام و نام خانوادگی", "نام کاربری", "رمز اولیه"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F3EEE4")
        cell.alignment = Alignment(horizontal="center")
    for item in created:
        sheet.append([item["code"], item["name"], item["code"], item["password"]])
    for column, width in zip("ABCD", (14, 30, 14, 16)):
        sheet.column_dimensions[column].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="portal-accounts.xlsx"'
    return response


def password_change(request):
    """تغییر رمز — هم برای کارکنان مالی، هم برای پرسنل در پرتال."""
    if not request.user.is_authenticated:
        return redirect("login")

    form = PasswordChangeForm(request.user, request.POST or None)
    for field in form.fields.values():
        field.widget.attrs["class"] = "input"

    if request.method == "POST" and form.is_valid():
        user = form.save()
        user.must_change_password = False
        user.save(update_fields=["must_change_password"])
        update_session_auth_hash(request, user)
        messages.success(request, "رمز عبور شما تغییر کرد.")
        return redirect("portal_home" if user.role == User.Role.EMPLOYEE else "dashboard")

    return render(
        request,
        "accounts/password_change.html",
        {"form": form, "forced": request.user.must_change_password},
    )
